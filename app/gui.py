"""
GST Reconciliation Tool — desktop GUI (Tkinter, stdlib only).

Layout:
  Login screen: PIN entry (existing install) or PIN setup (first run).
  Main window left panel  : client list (add / search / delete)
  Main window right panel : tabs for Upload, History, Reconciliation, Past Reports
"""
import os
import sys
import json
import ctypes
import subprocess
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta

import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FuncFormatter

from . import db, parsers, reconcile, report, service, security, licensing

# Reference palette (light mode) — see the dataviz skill's references/palette.md.
# Categorical slots are used in this fixed order (never cycled/generated) for the
# 6 fixed reconciliation categories; CHART_SEQUENTIAL is the single hue used for
# every magnitude/ranking chart (top suppliers, monthly trend).
CHART_CATEGORICAL = ["#2a78d6", "#1baf7a", "#eda100", "#008300", "#4a3aa7", "#e34948"]
CHART_SEQUENTIAL = "#2a78d6"
CHART_INK = "#0b0b0b"
CHART_MUTED = "#898781"
CHART_GRID = "#e1e0d9"
CHART_SURFACE = "#fcfcfb"


def format_inr(value, decimals=0):
    """Indian digit grouping (last 3 digits, then pairs: 1,23,45,678) rather than
    Western thousands-grouping — this is a GST tool for Indian accounting, and
    ₹5,425,630 (Western) reads oddly next to the lakh/crore figures accountants
    actually use, i.e. ₹54,25,630."""
    neg = value < 0
    value = abs(value)
    s = f"{value:,.{decimals}f}"
    int_part, _, dec_part = s.replace(",", "").partition(".")
    if len(int_part) > 3:
        last3 = int_part[-3:]
        rest = int_part[:-3]
        groups = []
        while len(rest) > 2:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.insert(0, rest)
        int_part = ",".join(groups) + "," + last3
    out = f"₹{int_part}" + (f".{dec_part}" if dec_part else "")
    return f"-{out}" if neg else out


def _style_axes(ax):
    """Shared chart chrome (recessive spines/gridlines/ticks) used by every
    matplotlib Axes in the app — the live Insights tab and the past-report
    viewer both draw the same three chart kinds and should look identical."""
    ax.set_facecolor(CHART_SURFACE)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(CHART_MUTED)
    ax.tick_params(colors=CHART_MUTED, labelsize=8)
    ax.title.set_color(CHART_INK)


def format_inr_short(value):
    """Compact lakh/crore notation for chart labels/axis ticks, e.g. 5425630 -> '₹54.3L'."""
    neg = value < 0
    value = abs(value)
    sign = "-" if neg else ""
    if value >= 1e7:
        return f"{sign}₹{value / 1e7:.2f}Cr"
    if value >= 1e5:
        return f"{sign}₹{value / 1e5:.1f}L"
    if value >= 1e3:
        return f"{sign}₹{value / 1e3:.1f}K"
    return f"{sign}₹{value:.0f}"

APP_TITLE = "GST Reconciliation Tool"
APP_VERSION = "1.1.4"

# User-facing release notes, newest first. Shown in-app via the "What's New"
# button (see ChangelogWindow) so the user has a reference without needing to
# read HANDOFF.md (which is a dev/session log, not meant for the end user).
CHANGELOG = [
    ("1.1.4", "Fixed: unable to set PIN on some displays", [
        "On some displays (high-DPI/scaled screens), the first-run 'set a PIN' screen cut "
        "off the Confirm PIN field and the button below the visible window, with no way to "
        "resize or scroll to reach them — blocking first-time setup entirely. The Activation "
        "screen and Add Client dialog had the same latent issue. All three now size themselves "
        "to fit their actual content correctly at any display scaling.",
    ]),
    ("1.1.3", "Installer: start-fresh option", [
        "The installer now offers an opt-in ‘Start fresh’ checkbox (unchecked by "
        "default) for setting up on a computer that has an older install's data still on "
        "it. When checked, previous clients/uploads/reports are archived aside (not "
        "permanently deleted) so a new PIN can be set up from scratch. Your activation key "
        "is never affected either way.",
    ]),
    ("1.1.2", "Note-matching fixes", [
        "Credit Note Register rows that are actually sales-side entries (a 'GST Sales' "
        "value present) are now excluded from purchase-side reconciliation, instead of "
        "showing up as false unmatched rows.",
        "GSTR-2B files that don't carry a portal generation date (some export variants "
        "omit it) can now have it entered manually — at upload time, or later via "
        "'Set Generation Date...' in the History tab.",
    ]),
    ("1.1.1", "Conflict details + What's New", [
        "Pending-conflict screens (both in the app and in the exported Excel report) now show "
        "the supplier/party name (Particulars), not just GSTIN and invoice/voucher number.",
        "Added this “What's New” viewer.",
    ]),
    ("1.1.0", "Credit/Debit Note reconciliation", [
        "Upload your Tally Credit Note Register and Debit Note Register, and reconcile them "
        "against the credit/debit notes in your GSTR-2B (the B2B-CDNR data) — the same way "
        "invoices are already reconciled against GSTR-2A/2B.",
        "New “Credit/Debit Notes” view under the Reconciliation tab.",
        "Uploading GSTR-2A/2B now also picks up credit/debit notes from the same file "
        "automatically — no extra upload step.",
        "Fixed: some Purchase Register uploads were undercounting IGST depending on how the "
        "IGST column was named in the export.",
        "Fixed: a database upgrade issue that could affect existing installations.",
    ]),
    ("1.0.7", "Report viewer", [
        "Double-click a past exported report to reopen it and see the same Insights/Details "
        "view, without re-running reconciliation against possibly-changed data.",
    ]),
    ("1.0.6", "Insights charts", [
        "New Insights sub-tab with KPI tiles and charts (category breakdown, monthly totals, "
        "top suppliers not yet in GSTR-2A/2B).",
        "All currency now shown in Indian digit grouping (₹54,25,630 style).",
        "App now launches maximized.",
    ]),
    ("1.0.5", "Cleaner look", [
        "Restyled the whole app for a cleaner, more readable look; tables now scroll instead "
        "of clipping long values.",
    ]),
    ("1.0.4", "Reconciliation preview fixes", [
        "Fixed some reconciliation categories showing blank/NaN values in the on-screen preview.",
        "History tables now show a Snapshot/Batch number for easier reference.",
    ]),
    ("1.0.3", "GSTR-2B support", [
        "GSTR-2B portal exports are now supported alongside GSTR-2A.",
        "Added Download/Delete for uploaded snapshots and Purchase Register batches.",
    ]),
    ("1.0.2", "High-DPI fix", [
        "Fixed garbled/blank button text on high-DPI displays.",
        "Upload buttons now support selecting multiple files at once.",
    ]),
    ("1.0.1", "Purchase Register upload fix", [
        "Fixed Purchase Register uploads being rejected with “No valid data rows were found.”",
    ]),
]

# Mirrors report.py's per-category column choices: the raw reconciliation
# DataFrames are outer-merge results carrying BOTH sides' columns (g2a_agg's
# come first since it's the merge's left side), so a naive "first 12 columns"
# preview shows all-NaN _2a columns for rows that only exist in the Purchase
# Register (and vice versa). Picking the side that's actually populated per
# category keeps the on-screen preview consistent with the exported report.
_CATEGORY_COLUMNS = {
    "value_tax_mismatches": (
        ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_no_pr", "invoice_date_2a",
         "invoice_value", "gross_total", "diff_value", "igst_2a", "igst_pr", "diff_igst",
         "cgst_2a", "cgst_pr", "diff_cgst", "sgst_2a", "sgst_pr", "diff_sgst"],
        {"gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No (2A)",
         "invoice_no_pr": "Invoice No (Books)", "invoice_date_2a": "Invoice Date",
         "invoice_value": "Invoice Value (2A)", "gross_total": "Gross Total (Books)",
         "diff_value": "Diff Value", "igst_2a": "IGST (2A)", "igst_pr": "IGST (Books)",
         "diff_igst": "Diff IGST", "cgst_2a": "CGST (2A)", "cgst_pr": "CGST (Books)",
         "diff_cgst": "Diff CGST", "sgst_2a": "SGST (2A)", "sgst_pr": "SGST (Books)", "diff_sgst": "Diff SGST"},
    ),
    "probable_matches": (
        ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_no_pr", "invoice_date_2a",
         "invoice_value", "gross_total"],
        {"gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No (2A)",
         "invoice_no_pr": "Invoice No (Books)", "invoice_date_2a": "Invoice Date",
         "invoice_value": "Invoice Value (2A)", "gross_total": "Gross Total (Books)"},
    ),
    "only_in_gstr2a": (
        ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_date_2a", "invoice_value",
         "taxable_value", "igst_2a", "cgst_2a", "sgst_2a", "itc_available", "period"],
        {"gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No",
         "invoice_date_2a": "Invoice Date", "invoice_value": "Invoice Value",
         "taxable_value": "Taxable Value", "igst_2a": "IGST", "cgst_2a": "CGST", "sgst_2a": "SGST",
         "itc_available": "ITC Available"},
    ),
    "only_in_purchase_register": (
        ["gstin_pr", "particulars", "invoice_no_pr", "invoice_date_pr", "gross_total",
         "igst_pr", "cgst_pr", "sgst_pr"],
        {"gstin_pr": "GSTIN", "particulars": "Supplier Name", "invoice_no_pr": "Invoice No",
         "invoice_date_pr": "Invoice Date", "gross_total": "Gross Total", "igst_pr": "IGST",
         "cgst_pr": "CGST", "sgst_pr": "SGST"},
    ),
    "matched_clean": (
        ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_date_2a", "invoice_value",
         "igst_2a", "cgst_2a", "sgst_2a", "itc_available"],
        {"gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No",
         "invoice_date_2a": "Invoice Date", "invoice_value": "Invoice Value", "igst_2a": "IGST",
         "cgst_2a": "CGST", "sgst_2a": "SGST", "itc_available": "ITC Available"},
    ),
    "multi_rate_reference": (
        ["period", "gstin", "supplier_name", "invoice_no", "invoice_date", "invoice_value",
         "rate", "taxable_value", "igst", "cgst", "sgst"],
        {"supplier_name": "Supplier Name"},
    ),
}

# Same idea as _CATEGORY_COLUMNS above, for reconcile.run_note_reconciliation()'s
# category DataFrames. Kept as a separate dict (not merged into _CATEGORY_COLUMNS)
# because the category names ('matched_clean', 'value_tax_mismatches', ...) are
# shared with the invoice side but the underlying columns differ (note_no/voucher_no
# instead of invoice_no, GSTR/Books instead of 2A/PR, etc.) — the note results live
# in their own dict (self._note_results) so there's no key collision at lookup time.
_NOTE_CATEGORY_COLUMNS = {
    "value_tax_mismatches": (
        ["gstin_gstr", "supplier_name", "note_no", "voucher_no", "note_type", "book_note_type",
         "note_date", "note_value", "gross_total", "diff_value", "igst_gstr", "igst_books", "diff_igst",
         "cgst_gstr", "cgst_books", "diff_cgst", "sgst_gstr", "sgst_books", "diff_sgst"],
        {"gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No (GSTR)",
         "voucher_no": "Voucher No (Books)", "note_type": "Note Type (GSTR)", "book_note_type": "Note Type (Books)",
         "note_date": "Note Date", "note_value": "Note Value (GSTR)", "gross_total": "Gross Total (Books)",
         "diff_value": "Diff Value", "igst_gstr": "IGST (GSTR)", "igst_books": "IGST (Books)",
         "diff_igst": "Diff IGST", "cgst_gstr": "CGST (GSTR)", "cgst_books": "CGST (Books)",
         "diff_cgst": "Diff CGST", "sgst_gstr": "SGST (GSTR)", "sgst_books": "SGST (Books)", "diff_sgst": "Diff SGST"},
    ),
    "probable_matches": (
        ["gstin_gstr", "supplier_name", "note_no", "voucher_no", "note_type", "book_note_type",
         "note_date", "note_value", "gross_total"],
        {"gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No (GSTR)",
         "voucher_no": "Voucher No (Books)", "note_type": "Note Type (GSTR)", "book_note_type": "Note Type (Books)",
         "note_date": "Note Date", "note_value": "Note Value (GSTR)", "gross_total": "Gross Total (Books)"},
    ),
    "only_in_gstr2b_cdnr": (
        ["gstin_gstr", "supplier_name", "note_no", "note_type", "note_date", "note_value",
         "taxable_value", "igst_gstr", "cgst_gstr", "sgst_gstr", "itc_available", "period"],
        {"gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No",
         "note_type": "Note Type", "note_date": "Note Date", "note_value": "Note Value",
         "taxable_value": "Taxable Value", "igst_gstr": "IGST", "cgst_gstr": "CGST", "sgst_gstr": "SGST",
         "itc_available": "ITC Available"},
    ),
    "only_in_note_register": (
        ["gstin_books", "particulars", "voucher_no", "book_note_type", "voucher_date",
         "gross_total", "igst_books", "cgst_books", "sgst_books"],
        {"gstin_books": "GSTIN", "particulars": "Supplier Name", "voucher_no": "Voucher No",
         "book_note_type": "Note Type (Books)", "voucher_date": "Voucher Date", "gross_total": "Gross Total",
         "igst_books": "IGST", "cgst_books": "CGST", "sgst_books": "SGST"},
    ),
    "matched_clean": (
        ["gstin_gstr", "supplier_name", "note_no", "note_type", "note_date", "note_value", "itc_available"],
        {"gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No",
         "note_type": "Note Type", "note_date": "Note Date", "note_value": "Note Value",
         "itc_available": "ITC Available"},
    ),
}


def _enable_dpi_awareness():
    """Without this, Windows bitmap-stretches the whole window on scaled displays,
    which mangles ttk widgets (buttons, tabs) rendered natively via the theme
    engine — their text comes out garbled/missing while plain Tk-drawn text is
    unaffected. Must run once, before any Tk root is created."""
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)  # PROCESS_SYSTEM_DPI_AWARE
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def _apply_dpi_scaling(root):
    """Once the process opts out of Windows' bitmap-stretching (see
    _enable_dpi_awareness), Tk must scale its own fonts/widgets to match the
    real display DPI or the UI renders tiny on high-DPI screens."""
    try:
        dpi = root.winfo_fpixels("1i")
        root.tk.call("tk", "scaling", dpi / 72.0)
    except Exception:
        pass


def _fit_window_to_content(window, min_w=0, min_h=0):
    """Sizes a fixed, non-resizable window to its ACTUAL required size instead
    of a hardcoded pixel guess. A hardcoded '440x300'-style geometry is a raw
    pixel count that _apply_dpi_scaling knows nothing about -- on a high-DPI
    display, scaling inflates every widget's real on-screen size, so content
    that fit at 100% scaling silently overflows past the bottom of a fixed,
    non-resizable window (buttons/fields below the fold, unreachable, with no
    scrollbar and no way to resize into view). Call this after packing every
    widget, in place of a hardcoded geometry() call, on any Tk/Toplevel using
    resizable(False, False) or resizable(False, True)."""
    window.update_idletasks()
    w = max(window.winfo_reqwidth(), min_w)
    h = max(window.winfo_reqheight(), min_h)
    window.geometry(f"{w}x{h}")


UI_FONT = "Segoe UI"


def _configure_style(root):
    """Single place defining every font/spacing/color choice, shared by all three
    top-level windows (Activation/Login/App each run their own Tk root and thus
    their own ttk.Style instance) so the look stays consistent and changes here
    apply everywhere instead of hunting down repeated inline font tuples."""
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    root.option_add("*Font", (UI_FONT, 10))
    style.configure(".", font=(UI_FONT, 10))
    style.configure("TButton", font=(UI_FONT, 10), padding=(10, 6))
    style.configure("TEntry", padding=4)
    style.configure("TCombobox", padding=4)
    style.configure("TNotebook.Tab", font=(UI_FONT, 10), padding=(18, 9))
    style.configure("Treeview", font=(UI_FONT, 9), rowheight=26)
    style.configure("Treeview.Heading", font=(UI_FONT, 9, "bold"))

    style.configure("Title.TLabel", font=(UI_FONT, 15, "bold"))
    style.configure("Header.TLabel", font=(UI_FONT, 13, "bold"))
    style.configure("Section.TLabel", font=(UI_FONT, 10, "bold"))
    style.configure("Muted.TLabel", foreground="#595959")
    style.configure("Error.TLabel", foreground="#9C0006")
    style.configure("Summary.TLabel", font=(UI_FONT, 10, "bold"), foreground="#1F4E78")
    return style


_COL_WIDTHS = [
    (("supplier", "particulars"), 250),
    (("file",), 220),
    (("uploaded",), 150),
    (("period",), 130),
    (("invoice no", "invoice_no"), 130),
    (("gstin",), 130),
    (("date",), 100),
    (("snapshot #", "batch #"), 80),
]
_COL_DEFAULT_WIDTH = 110


def _col_width(name):
    """Heuristic column width by header text, shared across every Treeview in
    the app — narrow for IDs/dates/amounts, wide for names/filenames — so long
    values (supplier names, filenames, timestamps) aren't clipped to a few
    characters, and long header text (e.g. 'Portal Generation Date') isn't
    clipped either."""
    key = name.lower()
    header_estimate = max(90, len(name) * 7 + 20)
    for keywords, width in _COL_WIDTHS:
        if any(k in key for k in keywords):
            return max(width, header_estimate)
    return max(_COL_DEFAULT_WIDTH, header_estimate)


def _make_scrollable_tree(parent, columns, height=10):
    """A ttk.Treeview with vertical + horizontal scrollbars wired up and column
    widths pre-sized via _col_width. Result sets here regularly exceed both the
    visible height (hundreds of rows) and width (a dozen+ columns for some
    reconciliation categories), so every Treeview in this app needs this same
    scaffolding — built once here instead of repeated per call site."""
    container = ttk.Frame(parent)
    tree = ttk.Treeview(container, columns=columns, show="headings", height=height)
    vsb = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    container.rowconfigure(0, weight=1)
    container.columnconfigure(0, weight=1)
    for c in columns:
        tree.heading(c, text=c)
        tree.column(c, width=_col_width(c), anchor="w")
    return container, tree


def _make_scrollable_frame(parent, bg=None):
    """A vertically scrollable area for content taller than the tab (e.g. several
    stacked chart panels) — a plain pack/grid layout has no scroll of its own, so
    without this, tall content just gets cut off by the window instead of being
    reachable by scrolling."""
    container = ttk.Frame(parent)
    canvas = tk.Canvas(container, highlightthickness=0, bg=bg or CHART_SURFACE)
    vsb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    inner = ttk.Frame(canvas)
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=vsb.set)
    canvas.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")
    canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
    return container, inner


def _bind_dynamic_wraplength(label, container, margin=24):
    """Keeps a label's wraplength in sync with its container's actual width. A
    static wraplength is wrong in either direction once a window can be resized
    or maximized: too small and long status lines wrap early even with plenty of
    room free (e.g. the reconciliation summary line); too large and it overflows
    a narrowed window instead of wrapping."""
    def _update(event=None):
        width = container.winfo_width() - margin
        if width > 100:
            label.configure(wraplength=width)
    container.bind("<Configure>", _update)
    label.after(50, _update)


def icon_path():
    """Location of the app icon, whether running from source or from a
    PyInstaller-frozen exe (whose bundled data lands in sys._MEIPASS)."""
    if getattr(sys, "frozen", False):
        base = sys._MEIPASS
    else:
        base = os.path.join(os.path.dirname(__file__), "..")
    path = os.path.join(base, "assets", "icon.ico")
    return path if os.path.exists(path) else None


def set_app_icon(window):
    path = icon_path()
    if path:
        try:
            window.iconbitmap(path)
        except tk.TclError:
            pass


def open_file(path):
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("Could not open file", str(e))


class ActivationScreen(tk.Tk):
    """Shown before anything else, on every machine that hasn't been activated
    yet. Displays this machine's fingerprint (to send to the vendor) and takes
    the activation key the vendor issues for it -- see app/licensing.py."""

    def __init__(self):
        super().__init__()
        _apply_dpi_scaling(self)
        _configure_style(self)
        self.title(APP_TITLE)
        self.resizable(False, False)
        set_app_icon(self)
        self.activated = False

        frame = ttk.Frame(self, padding=24)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Activate " + APP_TITLE, style="Title.TLabel").pack(anchor="w")
        ttk.Label(frame, text="This copy of the app hasn't been activated on this computer yet. "
                              "Send the machine ID below to whoever gave you this software; they'll "
                              "send back an activation key to paste in below.",
                  style="Muted.TLabel", wraplength=520, justify="left").pack(anchor="w", pady=(8, 18))

        ttk.Label(frame, text="This computer's Machine ID").pack(anchor="w")
        fp_row = ttk.Frame(frame)
        fp_row.pack(fill="x", pady=(3, 18))
        self.fp_var = tk.StringVar(value=licensing.current_fingerprint())
        fp_entry = ttk.Entry(fp_row, textvariable=self.fp_var, state="readonly")
        fp_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(fp_row, text="Copy", command=self._copy_fingerprint).pack(side="left", padx=(8, 0))

        ttk.Label(frame, text="Activation Key").pack(anchor="w")
        self.key_var = tk.StringVar()
        key_entry = ttk.Entry(frame, textvariable=self.key_var)
        key_entry.pack(fill="x", pady=(3, 12))
        key_entry.bind("<Return>", lambda e: self._submit())
        key_entry.focus_set()

        self.error_var = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.error_var, style="Error.TLabel", wraplength=520,
                  justify="left").pack(anchor="w", pady=(0, 10))

        ttk.Button(frame, text="Activate", command=self._submit).pack(anchor="e")

        _fit_window_to_content(self, min_w=580)

    def _copy_fingerprint(self):
        self.clipboard_clear()
        self.clipboard_append(self.fp_var.get())

    def _submit(self):
        key = self.key_var.get().strip()
        if not key:
            self.error_var.set("Paste the activation key you were given.")
            return
        try:
            licensing.activate(key)
        except licensing.LicenseError as e:
            self.error_var.set(str(e))
            return
        self.activated = True
        self.destroy()


class LoginScreen(tk.Tk):
    """Shown before anything else. Branches to PIN setup or PIN entry
    depending on db.needs_setup(), and only unblocks the main App once
    db.unlock_or_init() succeeds."""

    def __init__(self):
        super().__init__()
        _apply_dpi_scaling(self)
        _configure_style(self)
        self.title(APP_TITLE)
        self.resizable(False, False)
        set_app_icon(self)
        self.unlocked = False

        self.is_first_run = db.needs_setup()

        frame = ttk.Frame(self, padding=24)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=APP_TITLE, style="Title.TLabel").pack(anchor="w")

        if self.is_first_run:
            ttk.Label(frame, text="First run on this machine — set a PIN to protect your data.\n"
                                  "This PIN is required every time the app opens, and cannot be "
                                  "recovered if forgotten (see vendor support for a reset).",
                      style="Muted.TLabel", wraplength=390, justify="left").pack(anchor="w", pady=(8, 18))
            ttk.Label(frame, text="New PIN").pack(anchor="w")
            self.pin_var = tk.StringVar()
            ttk.Entry(frame, textvariable=self.pin_var, show="*").pack(fill="x", pady=(3, 12))
            ttk.Label(frame, text="Confirm PIN").pack(anchor="w")
            self.pin2_var = tk.StringVar()
            entry2 = ttk.Entry(frame, textvariable=self.pin2_var, show="*")
            entry2.pack(fill="x", pady=(3, 12))
            entry2.bind("<Return>", lambda e: self._submit())
            btn_text = "Set PIN and continue"
        else:
            ttk.Label(frame, text="Enter your PIN to unlock this client's data.",
                      style="Muted.TLabel", wraplength=390, justify="left").pack(anchor="w", pady=(8, 18))
            ttk.Label(frame, text="PIN").pack(anchor="w")
            self.pin_var = tk.StringVar()
            entry = ttk.Entry(frame, textvariable=self.pin_var, show="*")
            entry.pack(fill="x", pady=(3, 12))
            entry.bind("<Return>", lambda e: self._submit())
            entry.focus_set()
            btn_text = "Unlock"

        self.error_var = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.error_var, style="Error.TLabel",
                  wraplength=390, justify="left").pack(anchor="w", pady=(0, 10))

        ttk.Button(frame, text=btn_text, command=self._submit).pack(anchor="e")

        _fit_window_to_content(self, min_w=440)

    def _submit(self):
        pin = self.pin_var.get()
        if not pin:
            self.error_var.set("PIN cannot be empty.")
            return
        if self.is_first_run:
            if pin != self.pin2_var.get():
                self.error_var.set("PINs do not match.")
                return
            if len(pin) < 4:
                self.error_var.set("Use at least 4 characters.")
                return
        try:
            db.unlock_or_init(pin)
        except security.SecurityError:
            self.error_var.set("Incorrect PIN, or this data was created on a different computer.")
            return
        except Exception as e:
            self.error_var.set(f"Could not unlock: {e}")
            return
        self.unlocked = True
        self.destroy()


class ConflictResolutionDialog(tk.Toplevel):
    """Shows every Purchase Register (or Credit/Debit Note Register, via entity=
    'note') row awaiting an Overwrite/Ignore decision, stored value vs
    newly-uploaded value side by side, per HANDOFF's 'one real GUI gap'. Wired to
    db.list_pending_conflicts()/db.resolve_conflict() for entity='purchase', or
    the identical note_ variants for entity='note' -- the two entities use the
    same conflict-detection rules, so one dialog class serves both."""

    def __init__(self, parent, client_id, on_close=None, entity="purchase"):
        super().__init__(parent)
        self.client_id = client_id
        self.on_close = on_close
        self.entity = entity
        doc_word = "invoice" if entity == "purchase" else "credit/debit note"
        ref_field = "Invoice No" if entity == "purchase" else "Voucher No"
        self.title(f"Pending {'Purchase Register' if entity == 'purchase' else 'Credit/Debit Note Register'} Conflicts")
        self.geometry("1040x480")
        self.transient(parent)
        set_app_icon(self)

        ttk.Label(self, text=f"These {doc_word}s were re-uploaded with a DIFFERENT amount than what's "
                             "already stored. They are excluded from reconciliation until you resolve "
                             "each one — Overwrite uses the new upload, Ignore keeps the stored value.",
                  style="Muted.TLabel", wraplength=1000, justify="left", padding=10).pack(anchor="w")

        cols = (ref_field, "Particulars", "GSTIN", "Stored Gross", "New Gross", "Stored CGST", "New CGST",
                "Stored SGST", "New SGST", "Stored IGST", "New IGST")
        tree_container, self.tree = _make_scrollable_tree(self, cols, height=14)
        tree_container.pack(fill="both", expand=True, padx=10)

        btns = ttk.Frame(self, padding=10)
        btns.pack(fill="x")
        ttk.Button(btns, text="Overwrite Selected", command=lambda: self._resolve_selected("overwrite")).pack(side="left")
        ttk.Button(btns, text="Ignore Selected", command=lambda: self._resolve_selected("ignore")).pack(side="left", padx=6)
        ttk.Separator(btns, orient="vertical").pack(side="left", fill="y", padx=10)
        ttk.Button(btns, text="Overwrite All", command=lambda: self._resolve_all("overwrite")).pack(side="left")
        ttk.Button(btns, text="Ignore All", command=lambda: self._resolve_all("ignore")).pack(side="left", padx=6)
        ttk.Button(btns, text="Close", command=self._close).pack(side="right")

        self._rows = []
        self._refresh()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self.grab_set()

    def _refresh(self):
        self.tree.delete(*self.tree.get_children())
        ref_key = "invoice_no" if self.entity == "purchase" else "voucher_no"
        list_fn = db.list_pending_conflicts if self.entity == "purchase" else db.list_pending_note_conflicts
        self._rows = list_fn(self.client_id)
        for item in self._rows:
            p, s = item["pending"], item["stored"]
            self.tree.insert("", "end", iid=str(p["id"]), values=(
                p[ref_key], p["particulars"], p["gstin"],
                s["gross_total"] if s else "n/a", p["gross_total"],
                s["cgst"] if s else "n/a", p["cgst"],
                s["sgst"] if s else "n/a", p["sgst"],
                s["igst"] if s else "n/a", p["igst"],
            ))
        if not self._rows:
            messagebox.showinfo("No conflicts", "No pending conflicts remain.", parent=self)
            self._close()

    def _resolve_selected(self, action):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select one or more rows first.", parent=self)
            return
        resolve_fn = db.resolve_conflict if self.entity == "purchase" else db.resolve_note_conflict
        for iid in sel:
            resolve_fn(int(iid), action)
        self._refresh()

    def _resolve_all(self, action):
        if not self._rows:
            return
        if not messagebox.askyesno("Confirm", f"{action.capitalize()} all {len(self._rows)} pending conflicts?", parent=self):
            return
        resolve_all_fn = db.resolve_all_conflicts if self.entity == "purchase" else db.resolve_all_note_conflicts
        resolve_all_fn(self.client_id, action)
        self._refresh()

    def _close(self):
        self.grab_release()
        self.destroy()
        if self.on_close:
            self.on_close()


class GenerationDateDialog(tk.Toplevel):
    """Lets the user manually fill in a GSTR-2B snapshot's portal generation date
    when it couldn't be read from the file (some portal export variants omit the
    'Read me' sheet that normally carries it — not a parsing failure, the data
    just isn't in that file). One row per snapshot, so a multi-file upload where
    several files are missing it can be filled in a single dialog instead of one
    popup per file. Also reused by History's 'Set Generation Date' action for
    already-uploaded snapshots still sitting at n/a."""

    def __init__(self, parent, snapshots, on_save=None):
        """snapshots: list of dicts with keys snapshot_id, filename, current (str, may be '')."""
        super().__init__(parent)
        self.on_save = on_save
        self.title("Set Portal Generation Date")
        self.resizable(False, True)
        self.transient(parent)
        set_app_icon(self)

        ttk.Label(self, text="These file(s) don't include a portal generation date (the export "
                             "didn't carry a 'Read me' sheet). Enter it if you know it — format "
                             "DD/MM/YYYY, or use the quick buttons — or leave blank to keep showing n/a.",
                  style="Muted.TLabel", wraplength=600, justify="left", padding=10).pack(anchor="w")

        today_str = datetime.now().strftime("%d/%m/%Y")
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")

        rows_frame = ttk.Frame(self, padding=(10, 0))
        rows_frame.pack(fill="both", expand=True)
        self._entries = []
        for snap in snapshots:
            row = ttk.Frame(rows_frame)
            row.pack(fill="x", pady=4)
            ttk.Label(row, text=snap["filename"], width=28).pack(side="left")
            var = tk.StringVar(value=snap.get("current", "") or "")
            ttk.Entry(row, textvariable=var, width=12).pack(side="left", padx=(6, 6))
            ttk.Button(row, text="Today", width=8,
                       command=lambda v=var: v.set(today_str)).pack(side="left")
            ttk.Button(row, text="Yesterday", width=10,
                       command=lambda v=var: v.set(yesterday_str)).pack(side="left", padx=(4, 0))
            self._entries.append((snap["snapshot_id"], var))

        btns = ttk.Frame(self, padding=10)
        btns.pack(fill="x")
        ttk.Button(btns, text="Save", command=self._save).pack(side="right")
        ttk.Button(btns, text="Skip", command=self.destroy).pack(side="right", padx=(0, 6))

        _fit_window_to_content(self, min_w=640)
        self.update_idletasks()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        w, h = self.winfo_width(), self.winfo_height()
        x = px + max(0, (pw - w) // 2)
        y = py + max(0, (ph - h) // 3)
        self.geometry(f"+{x}+{y}")

        self.grab_set()

    def _save(self):
        import datetime as _dt
        bad = []
        for snapshot_id, var in self._entries:
            text = var.get().strip()
            if not text:
                continue
            try:
                _dt.datetime.strptime(text, "%d/%m/%Y")
            except ValueError:
                bad.append(text)
        if bad:
            messagebox.showwarning(
                "Invalid date", f"Couldn't understand: {', '.join(bad)}\nUse DD/MM/YYYY, e.g. 14/10/2025.",
                parent=self,
            )
            return
        for snapshot_id, var in self._entries:
            text = var.get().strip()
            if text:
                db.update_snapshot_generation_date(snapshot_id, text)
        self.grab_release()
        self.destroy()
        if self.on_save:
            self.on_save()


# Sheet names report.py writes, in the order the picker should offer them
# (Summary is deliberately excluded — it's freeform letterhead text, not a table).
_REPORT_SHEET_ORDER = ["Value_Tax_Mismatches", "Probable_Matches", "Only_In_GSTR2A",
                       "Only_In_PurchaseRegister", "Duplicates_GSTR2A", "MultiRate_Reference",
                       "Matched_Clean", "Pending_Conflicts"]


def _safe_sum(df, col):
    if df is None or col not in df.columns:
        return 0
    return pd.to_numeric(df[col], errors="coerce").fillna(0).sum()


def _safe_count(df, required_col=None):
    if df is None or (required_col is not None and required_col not in df.columns):
        return 0
    return len(df)


def _load_report_sheets(path):
    """Reads back report.py's per-category sheets as clean DataFrames.

    The header isn't always row 1: several sheets have a note row written
    above the real header (Probable_Matches always; Only_In_PurchaseRegister
    and Pending_Conflicts only in certain data-dependent cases, e.g. report.py
    calls ws.insert_rows(1) for Only_In_PurchaseRegister only when a
    'past_6_month_window' column is present). Rather than hardcode a per-sheet
    offset that would silently break if report.py's layout shifts again, the
    header row is located by content: a real header is several short text
    cells; a note is one long sentence alone in column A, and a blank spacer
    row has nothing at all — both fail the "is this a header row" test below.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for name in _REPORT_SHEET_ORDER:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=10, values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            non_null = [c for c in row if c not in (None, "")]
            if len(non_null) >= 3 and all(isinstance(c, str) and len(c) < 40 for c in non_null):
                header_idx = i
                break
        if header_idx is None:
            result[name] = pd.DataFrame()  # placeholder-message sheet, e.g. "No rows found..."
            continue
        header = [str(c) if c is not None else "" for c in rows[header_idx]]
        data_rows = [list(r[:len(header)]) for r in ws.iter_rows(min_row=header_idx + 2, values_only=True)
                     if any(c is not None for c in r)]
        result[name] = pd.DataFrame(data_rows, columns=header)
    return result


class ReportViewerWindow(tk.Toplevel):
    """Reopens an already-exported reconciliation report (.xlsx) in the same
    Category Details / Insights layout as a live run — since the original
    upload data may since have changed (new uploads, deletions), this reads
    back the frozen numbers from the saved file itself rather than re-running
    reconciliation, so it always matches exactly what was actually reported."""

    def __init__(self, parent, report_path):
        super().__init__(parent)
        self.report_path = report_path
        self.title(f"Report — {os.path.basename(report_path)}")
        self.geometry("1300x820")
        self.transient(parent)
        set_app_icon(self)

        try:
            self.sheets = _load_report_sheets(report_path)
        except Exception as e:
            ttk.Label(self, text=f"Could not read this report file:\n{e}",
                      style="Error.TLabel", padding=20).pack(anchor="w")
            return

        ttk.Label(self, text=os.path.basename(report_path), style="Header.TLabel",
                  padding=(12, 12, 12, 0)).pack(anchor="w")

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=12)
        details_tab = ttk.Frame(notebook, padding=(0, 10))
        insights_tab = ttk.Frame(notebook, padding=(0, 10))
        notebook.add(details_tab, text="Category Details")
        notebook.add(insights_tab, text="Insights")

        self._build_details(details_tab)
        self._build_insights(insights_tab)

    def _build_details(self, t):
        available = [s for s in _REPORT_SHEET_ORDER if s in self.sheets]
        filt = ttk.Frame(t)
        filt.pack(fill="x", pady=(0, 8))
        ttk.Label(filt, text="View category:").pack(side="left")
        self.sheet_var = tk.StringVar(value=available[0] if available else "")
        sheet_combo = ttk.Combobox(filt, textvariable=self.sheet_var, state="readonly",
                                    width=32, values=available)
        sheet_combo.pack(side="left", padx=8)
        sheet_combo.bind("<<ComboboxSelected>>", lambda e: self._render_sheet())

        result_container, self.result_tree = _make_scrollable_tree(t, (), height=20)
        result_container.pack(fill="both", expand=True)
        self._render_sheet()

    def _render_sheet(self):
        self.result_tree.delete(*self.result_tree.get_children())
        self.result_tree["columns"] = ()
        name = self.sheet_var.get()
        df = self.sheets.get(name)
        if df is None or len(df) == 0:
            self.result_tree["columns"] = ("info",)
            self.result_tree.heading("info", text="Result")
            self.result_tree.insert("", "end", values=("No rows in this category.",))
            return
        cols = list(df.columns)[:17]
        self.result_tree["columns"] = cols
        for c in cols:
            self.result_tree.heading(c, text=c)
            self.result_tree.column(c, width=_col_width(c), anchor="w")
        for _, row in df.head(500).iterrows():
            self.result_tree.insert("", "end", values=[row[c] for c in cols])

    def _build_insights(self, t):
        kpi_row = ttk.Frame(t)
        kpi_row.pack(fill="x", pady=(0, 16))
        matched_value = _safe_sum(self.sheets.get("Matched_Clean"), "Invoice Value")
        itc_at_risk = _safe_sum(self.sheets.get("Only_In_PurchaseRegister"), "Gross Total")
        mismatch_value = _safe_sum(self.sheets.get("Value_Tax_Mismatches"), "Diff_Value")
        if isinstance(mismatch_value, (int, float)):
            mismatch_value = abs(mismatch_value)
        for i, (label, val) in enumerate([
            ("Matched Value", matched_value),
            ("ITC at Risk (Only in Books)", itc_at_risk),
            ("Mismatch Value (abs)", mismatch_value),
        ]):
            tile = ttk.Frame(kpi_row, relief="solid", borderwidth=1, padding=14)
            tile.pack(side="left", fill="both", expand=True, padx=(0 if i == 0 else 10, 0))
            ttk.Label(tile, text=format_inr(val), font=(UI_FONT, 18, "bold")).pack(anchor="w")
            ttk.Label(tile, text=label, style="Muted.TLabel").pack(anchor="w", pady=(2, 0))

        scroll_container, charts_area = _make_scrollable_frame(t)
        scroll_container.pack(fill="both", expand=True)

        labels = ["Matched\nClean", "Value/Tax\nMismatches", "Probable\nMatches",
                  "Only in\nGSTR-2A", "Only in\nBooks", "Pending\nConflicts"]
        values = [
            _safe_count(self.sheets.get("Matched_Clean"), "Invoice No"),
            _safe_count(self.sheets.get("Value_Tax_Mismatches"), "GSTIN"),
            _safe_count(self.sheets.get("Probable_Matches"), "GSTIN"),
            _safe_count(self.sheets.get("Only_In_GSTR2A"), "GSTIN"),
            _safe_count(self.sheets.get("Only_In_PurchaseRegister"), "GSTIN"),
            _safe_count(self.sheets.get("Pending_Conflicts"), "GSTIN"),
        ]
        fig1 = Figure(figsize=(10.5, 3.2), dpi=96, facecolor=CHART_SURFACE)
        ax1 = fig1.add_subplot(111)
        bars = ax1.bar(labels, values, color=CHART_CATEGORICAL, width=0.6)
        for bar, val in zip(bars, values):
            ax1.annotate(str(val), (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                         ha="center", va="bottom", fontsize=8, color=CHART_INK)
        ax1.set_title("Reconciliation Categories", fontsize=10, fontweight="bold", loc="left")
        ax1.grid(axis="y", color=CHART_GRID, linewidth=0.8)
        ax1.set_axisbelow(True)
        _style_axes(ax1)
        fig1.tight_layout()
        canvas1 = FigureCanvasTkAgg(fig1, master=charts_area)
        canvas1.get_tk_widget().pack(fill="both", expand=True, pady=(0, 16))
        canvas1.draw()

        fig2 = Figure(figsize=(10.5, 4.6), dpi=96, facecolor=CHART_SURFACE)
        ax2 = fig2.add_subplot(111)
        only_pr = self.sheets.get("Only_In_PurchaseRegister")
        if only_pr is not None and "Supplier Name" in only_pr.columns and "Gross Total" in only_pr.columns:
            top = only_pr.groupby("Supplier Name")["Gross Total"].sum().sort_values(ascending=False).head(8)
            top = top.sort_values(ascending=True)
            names = [n if len(str(n)) <= 24 else str(n)[:21] + "…" for n in top.index]
            ax2.barh(names, top.values, color=CHART_SEQUENTIAL, height=0.5)
            ax2.margins(x=0.16)
            for i, val in enumerate(top.values):
                ax2.annotate(format_inr_short(val), (val, i), ha="left", va="center", fontsize=8,
                             color=CHART_INK, xytext=(6, 0), textcoords="offset points")
        else:
            ax2.text(0.5, 0.5, "Nothing only in the Purchase Register in this report", ha="center",
                     va="center", color=CHART_MUTED, fontsize=9, transform=ax2.transAxes)
        ax2.set_title("Top Suppliers Not Yet in GSTR-2A/2B (ITC at Risk)", fontsize=10,
                       fontweight="bold", loc="left")
        ax2.grid(axis="x", color=CHART_GRID, linewidth=0.8)
        ax2.set_axisbelow(True)
        _style_axes(ax2)
        fig2.tight_layout()
        canvas2 = FigureCanvasTkAgg(fig2, master=charts_area)
        canvas2.get_tk_widget().pack(fill="both", expand=True)
        canvas2.draw()


class ChangelogWindow(tk.Toplevel):
    """Read-only 'What's New' viewer over the CHANGELOG constant above — lets the
    user see release notes without leaving the app or hunting for a GitHub page."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title(f"What's New — {APP_TITLE}")
        self.geometry("640x560")
        self.transient(parent)
        set_app_icon(self)

        ttk.Label(self, text=f"{APP_TITLE}  —  currently v{APP_VERSION}",
                  style="Header.TLabel", padding=(14, 12, 14, 4)).pack(anchor="w")

        text_frame = ttk.Frame(self, padding=(14, 0, 14, 14))
        text_frame.pack(fill="both", expand=True)
        text = tk.Text(text_frame, wrap="word", state="normal", relief="flat",
                        bg="#FFFFFF", font=(UI_FONT, 10), padx=8, pady=8)
        vsb = ttk.Scrollbar(text_frame, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=vsb.set)
        text.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        text.tag_configure("version", font=(UI_FONT, 12, "bold"), foreground="#1F4E78", spacing3=4)
        text.tag_configure("bullet", font=(UI_FONT, 10), lmargin1=14, lmargin2=28, spacing1=2, spacing3=6)
        for version, title, notes in CHANGELOG:
            text.insert("end", f"v{version} — {title}\n", "version")
            for note in notes:
                text.insert("end", f"• {note}\n", "bullet")
            text.insert("end", "\n")
        text.configure(state="disabled")

        ttk.Button(self, text="Close", command=self.destroy).pack(pady=(0, 12))
        self.grab_set()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        _apply_dpi_scaling(self)
        self.title(f"{APP_TITLE}  v{APP_VERSION}")
        self.geometry("1360x800")
        self.minsize(1080, 660)
        set_app_icon(self)
        try:
            self.state("zoomed")  # start maximized; falls back to the geometry above if unsupported
        except tk.TclError:
            pass

        self.current_client_id = None
        self._results = None
        self._current_fy_meta = None

        self._build_layout()
        self._refresh_client_list()

    # ---------------- layout ----------------

    def _build_layout(self):
        _configure_style(self)

        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # ---- left: client panel ----
        left = ttk.Frame(root, width=260)
        left.pack(side="left", fill="y", padx=(0, 12))
        left.pack_propagate(False)

        ttk.Label(left, text="Clients", style="Header.TLabel").pack(anchor="w", pady=(0, 6))

        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(left, textvariable=self.search_var)
        search_entry.pack(fill="x", pady=(0, 6))
        search_entry.bind("<KeyRelease>", lambda e: self._refresh_client_list())

        self.client_listbox = tk.Listbox(left, height=25, exportselection=False,
                                          font=(UI_FONT, 10), activestyle="none",
                                          relief="solid", borderwidth=1,
                                          highlightthickness=0)
        self.client_listbox.pack(fill="both", expand=True)
        self.client_listbox.bind("<<ListboxSelect>>", self._on_client_selected)

        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=(8, 0))
        ttk.Button(btns, text="+ Add Client", command=self._add_client_dialog).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Delete", command=self._delete_client).pack(side="left", expand=True, fill="x")

        # ---- right: client detail ----
        right = ttk.Frame(root)
        right.pack(side="left", fill="both", expand=True)

        header_row = ttk.Frame(right)
        header_row.pack(fill="x", pady=(0, 8))
        self.header_var = tk.StringVar(value="Select or add a client to begin")
        ttk.Label(header_row, textvariable=self.header_var, style="Title.TLabel").pack(side="left")
        ttk.Button(header_row, text="What's New", command=lambda: ChangelogWindow(self)).pack(side="left", padx=(14, 0))
        self.note_conflicts_btn = ttk.Button(header_row, text="Resolve Note Conflicts",
                                              command=self._open_note_conflicts_dialog, state="disabled")
        self.note_conflicts_btn.pack(side="right")
        self.conflicts_btn = ttk.Button(header_row, text="Resolve Purchase Conflicts",
                                         command=self._open_conflicts_dialog, state="disabled")
        self.conflicts_btn.pack(side="right", padx=(0, 8))

        self.notebook = ttk.Notebook(right)
        self.notebook.pack(fill="both", expand=True)

        self.tab_upload = ttk.Frame(self.notebook, padding=12)
        self.tab_history = ttk.Frame(self.notebook, padding=12)
        self.tab_recon = ttk.Frame(self.notebook, padding=12)
        self.tab_reports = ttk.Frame(self.notebook, padding=12)

        self.notebook.add(self.tab_upload, text="Upload Data")
        self.notebook.add(self.tab_history, text="History")
        self.notebook.add(self.tab_recon, text="Reconciliation")
        self.notebook.add(self.tab_reports, text="Past Reports")

        self._build_upload_tab()
        self._build_history_tab()
        self._build_recon_tab()
        self._build_reports_tab()

        self._set_tabs_enabled(False)

    def _set_tabs_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        for i in range(1, 4):
            self.notebook.tab(i, state=state if enabled else "disabled")
        for w in (self.upload_g2a_btn, self.upload_pr_btn, self.upload_cn_btn, self.upload_dn_btn):
            w.configure(state=state)
        self.conflicts_btn.configure(state=state)
        self.note_conflicts_btn.configure(state=state)

    # ---------------- client list ----------------

    def _refresh_client_list(self):
        self.client_listbox.delete(0, "end")
        query = self.search_var.get().strip().lower()
        self._clients_cache = [c for c in db.list_clients() if query in c["name"].lower()]
        for c in self._clients_cache:
            self.client_listbox.insert("end", f"{c['name']}" + (f"  ({c['gstin']})" if c["gstin"] else ""))

    def _on_client_selected(self, event):
        sel = self.client_listbox.curselection()
        if not sel:
            return
        client = self._clients_cache[sel[0]]
        self.current_client_id = client["client_id"]
        self.header_var.set(f"{client['name']}" + (f"   |   GSTIN {client['gstin']}" if client["gstin"] else ""))
        self._set_tabs_enabled(True)
        self._refresh_history_tab()
        self._refresh_recon_tab()
        self._refresh_reports_tab()

    def _add_client_dialog(self):
        dlg = tk.Toplevel(self)
        dlg.title("Add Client")
        dlg.resizable(False, False)
        dlg.transient(self)
        set_app_icon(dlg)
        ttk.Label(dlg, text="Client / Business Name*").pack(anchor="w", padx=16, pady=(16, 0))
        name_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=name_var).pack(fill="x", padx=16, pady=(3, 0))
        ttk.Label(dlg, text="GSTIN (optional)").pack(anchor="w", padx=16, pady=(10, 0))
        gstin_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=gstin_var).pack(fill="x", padx=16, pady=(3, 0))

        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("Name required", "Please enter a client name.")
                return
            try:
                db.add_client(name, gstin_var.get().strip())
            except Exception as e:
                messagebox.showerror("Could not add client", f"A client with this name may already exist.\n\n{e}")
                return
            dlg.destroy()
            self._refresh_client_list()

        ttk.Button(dlg, text="Save", command=save).pack(pady=14)
        _fit_window_to_content(dlg, min_w=400)
        dlg.grab_set()

    def _delete_client(self):
        sel = self.client_listbox.curselection()
        if not sel:
            return
        client = self._clients_cache[sel[0]]
        if messagebox.askyesno(
            "Delete client",
            f"This permanently deletes ALL stored data for '{client['name']}' "
            f"(GSTR-2A snapshots, purchase entries, and reconciliation history).\n\nContinue?",
        ):
            db.delete_client(client["client_id"])
            self.current_client_id = None
            self.header_var.set("Select or add a client to begin")
            self._set_tabs_enabled(False)
            self._refresh_client_list()

    # ---------------- Pending conflicts dialog ----------------

    def _open_conflicts_dialog(self):
        if not self.current_client_id:
            return
        pending = db.list_pending_conflicts(self.current_client_id)
        if not pending:
            messagebox.showinfo("No conflicts", "There are no pending Purchase Register conflicts for this client.")
            return
        ConflictResolutionDialog(self, self.current_client_id, on_close=self._on_conflicts_closed, entity="purchase")

    def _open_note_conflicts_dialog(self):
        if not self.current_client_id:
            return
        pending = db.list_pending_note_conflicts(self.current_client_id)
        if not pending:
            messagebox.showinfo("No conflicts", "There are no pending Credit/Debit Note Register conflicts for this client.")
            return
        ConflictResolutionDialog(self, self.current_client_id, on_close=self._on_conflicts_closed, entity="note")

    def _on_conflicts_closed(self):
        self._refresh_history_tab()
        self._refresh_recon_tab()

    # ---------------- Upload tab ----------------

    def _build_upload_tab(self):
        t = self.tab_upload
        ttk.Label(t, text="Step 1 — Upload the GSTR-2A/2B file downloaded from the GST portal",
                  style="Section.TLabel").pack(anchor="w")
        lbl1 = ttk.Label(t, text="Each upload is kept as a dated snapshot — nothing is overwritten, "
                          "so you can track suppliers filing late over the following months.",
                  style="Muted.TLabel", justify="left")
        lbl1.pack(anchor="w", pady=(2, 8))
        _bind_dynamic_wraplength(lbl1, t)
        self.upload_g2a_btn = ttk.Button(t, text="Select GSTR-2A/2B file(s) (.xls/.xlsx)...",
                                          command=self._upload_gstr2a, state="disabled")
        self.upload_g2a_btn.pack(anchor="w", pady=(0, 18))

        ttk.Label(t, text="Step 2 — Upload the Purchase Register (monthly file, or a full year at once)",
                  style="Section.TLabel").pack(anchor="w")
        lbl2 = ttk.Label(t, text="Upload each month's file as it becomes available, or a full-year file — rows "
                          "are auto-tagged by month/FY. Anything already stored is automatically skipped, "
                          "so it's safe to re-upload the same file by mistake. Re-uploading the same "
                          "invoice with a DIFFERENT amount is held as a pending conflict for review "
                          "(see the 'Resolve Purchase Conflicts' button above).",
                  style="Muted.TLabel", justify="left")
        lbl2.pack(anchor="w", pady=(2, 8))
        _bind_dynamic_wraplength(lbl2, t)
        self.upload_pr_btn = ttk.Button(t, text="Select Purchase Register file(s) (.xls/.xlsx)...",
                                         command=self._upload_purchase, state="disabled")
        self.upload_pr_btn.pack(anchor="w", pady=(0, 18))

        ttk.Label(t, text="Step 3 — Upload the Credit Note Register (from Tally)",
                  style="Section.TLabel").pack(anchor="w")
        lbl3 = ttk.Label(t, text="Same conflict-detection rules as Purchase Register: identical re-uploads "
                          "are skipped, a re-upload of the same voucher with a DIFFERENT amount is held as "
                          "a pending conflict (see 'Resolve Note Conflicts' button above).",
                  style="Muted.TLabel", justify="left")
        lbl3.pack(anchor="w", pady=(2, 8))
        _bind_dynamic_wraplength(lbl3, t)
        self.upload_cn_btn = ttk.Button(t, text="Select Credit Note Register file(s) (.xls/.xlsx)...",
                                         command=lambda: self._upload_note("credit"), state="disabled")
        self.upload_cn_btn.pack(anchor="w", pady=(0, 18))

        ttk.Label(t, text="Step 4 — Upload the Debit Note Register (from Tally)",
                  style="Section.TLabel").pack(anchor="w")
        lbl4 = ttk.Label(t, text="Same conflict-detection rules as Purchase Register.",
                  style="Muted.TLabel", justify="left")
        lbl4.pack(anchor="w", pady=(2, 8))
        _bind_dynamic_wraplength(lbl4, t)
        self.upload_dn_btn = ttk.Button(t, text="Select Debit Note Register file(s) (.xls/.xlsx)...",
                                         command=lambda: self._upload_note("debit"), state="disabled")
        self.upload_dn_btn.pack(anchor="w", pady=(0, 18))

        ttk.Separator(t).pack(fill="x", pady=10)
        self.upload_log = tk.Text(t, height=14, wrap="word", state="disabled",
                                   bg="#F7F7F7", relief="flat", font=(UI_FONT, 9),
                                   padx=8, pady=6)
        self.upload_log.pack(fill="both", expand=True)

    def _log(self, widget, msg):
        widget.configure(state="normal")
        widget.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        widget.see("end")
        widget.configure(state="disabled")

    def _upload_gstr2a(self):
        paths = filedialog.askopenfilenames(
            title="Select GSTR-2A file(s)",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if not paths:
            return
        total_rows = 0
        total_notes = 0
        all_periods = []
        errors = []
        missing_gen_date = []
        for path in paths:
            try:
                df, gen_date = parsers.parse_gstr2a(path)
                snap_id = db.add_gstr2a_snapshot(self.current_client_id, os.path.basename(path), gen_date, df)
                periods = sorted(df["Period"].unique())
                all_periods.extend(periods)
                total_rows += len(df)
                note_msg = ""
                cdnr_df, _ = parsers.parse_gstr2b_cdnr(path)
                if len(cdnr_df):
                    db.add_gstr2b_cdnr_notes(self.current_client_id, snap_id, cdnr_df)
                    total_notes += len(cdnr_df)
                    note_msg = f", {len(cdnr_df)} credit/debit note rows (B2B-CDNR)"
                self._log(self.upload_log,
                           f"{os.path.basename(path)}: {len(df)} invoice rows{note_msg}, periods {periods[0]}–{periods[-1]}, "
                           f"portal generation date {gen_date or 'n/a'}. Saved as snapshot #{snap_id}.")
                if not gen_date:
                    missing_gen_date.append({"snapshot_id": snap_id, "filename": os.path.basename(path), "current": ""})
            except parsers.ParseError as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}")
            except Exception as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}\n{traceback.format_exc()[-800:]}")

        if all_periods:
            msg = (f"{len(paths)} file(s) selected, {total_rows} invoice rows and {total_notes} credit/debit "
                   f"note rows saved as new snapshots.\n\nPeriods covered: {min(all_periods)} to {max(all_periods)}")
        else:
            msg = "No files were successfully processed."
        if errors:
            msg += "\n\nThe following file(s) failed:\n" + "\n".join(errors)
            messagebox.showwarning("GSTR-2A uploaded with errors", msg)
        else:
            messagebox.showinfo("GSTR-2A uploaded", msg)
        self._refresh_history_tab()
        self._refresh_recon_tab()
        if missing_gen_date:
            GenerationDateDialog(self, missing_gen_date, on_save=self._refresh_history_tab)

    def _upload_purchase(self):
        paths = filedialog.askopenfilenames(
            title="Select Purchase Register file(s)",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if not paths:
            return
        total_new = total_skipped = total_conflicts = 0
        errors = []
        for path in paths:
            try:
                entries = parsers.parse_purchase_register(path)
                result = db.add_purchase_batch(self.current_client_id, os.path.basename(path), entries)
                self._log(self.upload_log,
                           f"{os.path.basename(path)}: {result['total_in_file']} rows read → "
                           f"{result['new_rows']} new, {result['duplicate_rows_skipped']} already stored (skipped), "
                           f"{result['pending_conflicts']} flagged as amount conflicts for review.")
                total_new += result["new_rows"]
                total_skipped += result["duplicate_rows_skipped"]
                total_conflicts += result["pending_conflicts"]
            except parsers.ParseError as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}")
            except Exception as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}\n{traceback.format_exc()[-800:]}")

        msg = (f"{len(paths)} file(s) selected.\n"
               f"{total_new} new entries added.\n"
               f"{total_skipped} rows were already in the system and skipped.\n")
        if total_conflicts:
            msg += (f"\n{total_conflicts} entries have the same invoice number as one already "
                    f"stored but a DIFFERENT amount. Use 'Resolve Purchase Conflicts' above before "
                    f"relying on totals — these are excluded from reconciliation until resolved.")
        if errors:
            msg += "\n\nThe following file(s) failed:\n" + "\n".join(errors)
            messagebox.showwarning("Purchase Register uploaded with errors", msg)
        else:
            messagebox.showinfo("Purchase Register uploaded", msg)
        self._refresh_history_tab()
        self._refresh_recon_tab()

    def _upload_note(self, note_type):
        label = "Credit Note" if note_type == "credit" else "Debit Note"
        paths = filedialog.askopenfilenames(
            title=f"Select {label} Register file(s)",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if not paths:
            return
        total_new = total_skipped = total_conflicts = 0
        errors = []
        for path in paths:
            try:
                entries = parsers.parse_note_register(path, note_type)
                result = db.add_note_batch(self.current_client_id, note_type, os.path.basename(path), entries)
                self._log(self.upload_log,
                           f"{os.path.basename(path)}: {result['total_in_file']} rows read → "
                           f"{result['new_rows']} new, {result['duplicate_rows_skipped']} already stored (skipped), "
                           f"{result['pending_conflicts']} flagged as amount conflicts for review.")
                total_new += result["new_rows"]
                total_skipped += result["duplicate_rows_skipped"]
                total_conflicts += result["pending_conflicts"]
            except parsers.ParseError as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}")
            except Exception as e:
                errors.append(f"{os.path.basename(path)}: {e}")
                self._log(self.upload_log, f"{os.path.basename(path)}: FAILED — {e}\n{traceback.format_exc()[-800:]}")

        msg = (f"{len(paths)} file(s) selected.\n"
               f"{total_new} new entries added.\n"
               f"{total_skipped} rows were already in the system and skipped.\n")
        if total_conflicts:
            msg += (f"\n{total_conflicts} entries have the same voucher number as one already "
                    f"stored but a DIFFERENT amount. Use 'Resolve Note Conflicts' above before "
                    f"relying on totals — these are excluded from reconciliation until resolved.")
        if errors:
            msg += "\n\nThe following file(s) failed:\n" + "\n".join(errors)
            messagebox.showwarning(f"{label} Register uploaded with errors", msg)
        else:
            messagebox.showinfo(f"{label} Register uploaded", msg)
        self._refresh_history_tab()
        self._refresh_recon_tab()

    # ---------------- History tab ----------------

    def _build_history_tab(self):
        t = self.tab_history
        ttk.Label(t, text="GSTR-2A/2B Snapshots", style="Section.TLabel").pack(anchor="w", pady=(0, 6))
        cols = ("Snapshot #", "Uploaded", "File", "Portal Generation Date", "Periods", "Invoices")
        snap_container, self.snap_tree = _make_scrollable_tree(t, cols, height=8)
        snap_container.pack(fill="x", pady=(0, 8))
        snap_btns = ttk.Frame(t)
        snap_btns.pack(fill="x", pady=(0, 20))
        ttk.Button(snap_btns, text="Download Selected...", command=self._download_snapshot).pack(side="left")
        ttk.Button(snap_btns, text="Delete Selected", command=self._delete_snapshot).pack(side="left", padx=8)
        ttk.Button(snap_btns, text="Set Generation Date...", command=self._set_generation_date).pack(side="left")

        ttk.Label(t, text="Purchase Register Upload Batches", style="Section.TLabel").pack(anchor="w", pady=(0, 6))
        cols2 = ("Batch #", "Uploaded", "File", "Rows in File", "New", "Skipped (dup)", "Pending Conflicts")
        batch_container, self.batch_tree = _make_scrollable_tree(t, cols2, height=10)
        batch_container.pack(fill="both", expand=True, pady=(0, 8))
        batch_btns = ttk.Frame(t)
        batch_btns.pack(fill="x", pady=(0, 20))
        ttk.Button(batch_btns, text="Download Selected...", command=self._download_batch).pack(side="left")
        ttk.Button(batch_btns, text="Delete Selected", command=self._delete_batch).pack(side="left", padx=8)

        ttk.Label(t, text="Credit/Debit Note Register Upload Batches", style="Section.TLabel").pack(anchor="w", pady=(0, 6))
        cols3 = ("Batch #", "Type", "Uploaded", "File", "Rows in File", "New", "Skipped (dup)", "Pending Conflicts")
        note_batch_container, self.note_batch_tree = _make_scrollable_tree(t, cols3, height=10)
        note_batch_container.pack(fill="both", expand=True, pady=(0, 8))
        note_batch_btns = ttk.Frame(t)
        note_batch_btns.pack(fill="x")
        ttk.Button(note_batch_btns, text="Download Selected...", command=self._download_note_batch).pack(side="left")
        ttk.Button(note_batch_btns, text="Delete Selected", command=self._delete_note_batch).pack(side="left", padx=8)

    def _refresh_history_tab(self):
        if not self.current_client_id:
            return
        self.snap_tree.delete(*self.snap_tree.get_children())
        for s in db.list_snapshots(self.current_client_id):
            self.snap_tree.insert("", "end", iid=str(s["snapshot_id"]), values=(
                s["snapshot_id"], s["uploaded_at"], s["source_filename"], s["generation_date"] or "n/a",
                f"{s['period_min']}–{s['period_max']}", s["row_count"],
            ))
        self.batch_tree.delete(*self.batch_tree.get_children())
        for b in db.list_purchase_batches(self.current_client_id):
            self.batch_tree.insert("", "end", iid=str(b["batch_id"]), values=(
                b["batch_id"], b["uploaded_at"], b["source_filename"], b["row_count"], b["new_rows"],
                b["duplicate_rows_skipped"], b["conflict_rows"],
            ))
        self.note_batch_tree.delete(*self.note_batch_tree.get_children())
        for b in db.list_note_batches(self.current_client_id):
            self.note_batch_tree.insert("", "end", iid=str(b["batch_id"]), values=(
                b["batch_id"], "Credit Note" if b["note_type"] == "credit" else "Debit Note",
                b["uploaded_at"], b["source_filename"], b["row_count"], b["new_rows"],
                b["duplicate_rows_skipped"], b["conflict_rows"],
            ))

    def _download_snapshot(self):
        sel = self.snap_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select a snapshot first.")
            return
        snapshot_id = int(sel[0])
        default_name = f"GSTR2A_2B_snapshot_{snapshot_id}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save snapshot as", initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            service.export_gstr2a_snapshot(snapshot_id, path)
        except Exception as e:
            messagebox.showerror("Download failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return
        if messagebox.askyesno("Saved", f"Saved to:\n{path}\n\nOpen it now?"):
            open_file(path)

    def _delete_snapshot(self):
        sel = self.snap_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select a snapshot first.")
            return
        snapshot_id = int(sel[0])
        if not messagebox.askyesno(
            "Delete snapshot",
            "This removes this snapshot and all its invoice rows from this client's data. "
            "This cannot be undone (though you can re-upload the same file afterwards). Continue?",
        ):
            return
        db.delete_gstr2a_snapshot(snapshot_id)
        self._refresh_history_tab()
        self._refresh_recon_tab()

    def _set_generation_date(self):
        sel = self.snap_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select one or more snapshots first.")
            return
        selected_ids = {int(iid) for iid in sel}
        snaps = [s for s in db.list_snapshots(self.current_client_id) if s["snapshot_id"] in selected_ids]
        entries = [{"snapshot_id": s["snapshot_id"], "filename": s["source_filename"],
                    "current": s["generation_date"] or ""} for s in snaps]
        GenerationDateDialog(self, entries, on_save=self._refresh_history_tab)

    def _download_batch(self):
        sel = self.batch_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select an upload batch first.")
            return
        batch_id = int(sel[0])
        default_name = f"PurchaseRegister_batch_{batch_id}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save batch as", initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            service.export_purchase_batch(batch_id, path)
        except Exception as e:
            messagebox.showerror("Download failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return
        if messagebox.askyesno("Saved", f"Saved to:\n{path}\n\nOpen it now?"):
            open_file(path)

    def _delete_batch(self):
        sel = self.batch_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select an upload batch first.")
            return
        batch_id = int(sel[0])
        if not messagebox.askyesno(
            "Delete upload batch",
            "This removes every Purchase Register row this upload added from this client's data. "
            "This cannot be undone (though you can re-upload the same file afterwards). Continue?",
        ):
            return
        db.delete_purchase_batch(batch_id)
        self._refresh_history_tab()
        self._refresh_recon_tab()

    def _download_note_batch(self):
        sel = self.note_batch_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select an upload batch first.")
            return
        batch_id = int(sel[0])
        default_name = f"NoteRegister_batch_{batch_id}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save batch as", initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            service.export_note_batch(batch_id, path)
        except Exception as e:
            messagebox.showerror("Download failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return
        if messagebox.askyesno("Saved", f"Saved to:\n{path}\n\nOpen it now?"):
            open_file(path)

    def _delete_note_batch(self):
        sel = self.note_batch_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select an upload batch first.")
            return
        batch_id = int(sel[0])
        if not messagebox.askyesno(
            "Delete upload batch",
            "This removes every Credit/Debit Note Register row this upload added from this client's data. "
            "This cannot be undone (though you can re-upload the same file afterwards). Continue?",
        ):
            return
        db.delete_note_batch(batch_id)
        self._refresh_history_tab()
        self._refresh_recon_tab()

    # ---------------- Reconciliation tab ----------------

    def _build_recon_tab(self):
        t = self.tab_recon
        top = ttk.Frame(t)
        top.pack(fill="x", pady=(0, 8))
        ttk.Label(top, text="Financial Year:").pack(side="left")
        self.fy_var = tk.StringVar()
        self.fy_combo = ttk.Combobox(top, textvariable=self.fy_var, width=14, state="readonly")
        self.fy_combo.pack(side="left", padx=8)
        ttk.Button(top, text="Run Reconciliation", command=self._run_reconciliation).pack(side="left", padx=8)
        ttk.Label(top, text="GSTR-2A data used is automatically merged across every snapshot "
                            "uploaded for this client — no manual selection needed.",
                  style="Muted.TLabel", wraplength=560, justify="left").pack(side="left", padx=(16, 0))

        month_row = ttk.Frame(t)
        month_row.pack(fill="x", pady=(0, 14))
        ttk.Label(month_row, text="Export one month of Purchase Register data:").pack(side="left")
        self.month_var = tk.StringVar()
        self.month_combo = ttk.Combobox(month_row, textvariable=self.month_var, width=14, state="readonly")
        self.month_combo.pack(side="left", padx=8)
        ttk.Button(month_row, text="Export Month to Excel...", command=self._export_month).pack(side="left")

        self.recon_summary_var = tk.StringVar(value="")
        summary_lbl = ttk.Label(t, textvariable=self.recon_summary_var, style="Summary.TLabel", justify="left")
        summary_lbl.pack(anchor="w", pady=(0, 3))
        _bind_dynamic_wraplength(summary_lbl, t)
        self.recon_cutoff_var = tk.StringVar(value="")
        cutoff_lbl = ttk.Label(t, textvariable=self.recon_cutoff_var, justify="left")
        cutoff_lbl.pack(anchor="w", pady=(0, 10))
        _bind_dynamic_wraplength(cutoff_lbl, t)

        # Category detail (table) and Insights (KPIs/charts) share the FY controls
        # above but are otherwise two different views onto the same reconciliation
        # run, so they live as inner sub-tabs of Reconciliation rather than a
        # separate top-level tab.
        self.recon_inner_notebook = ttk.Notebook(t)
        self.recon_inner_notebook.pack(fill="both", expand=True)
        self.recon_subtab_details = ttk.Frame(self.recon_inner_notebook, padding=(0, 10))
        self.recon_subtab_insights = ttk.Frame(self.recon_inner_notebook, padding=(0, 10))
        self.recon_subtab_notes = ttk.Frame(self.recon_inner_notebook, padding=(0, 10))
        self.recon_inner_notebook.add(self.recon_subtab_details, text="Category Details")
        self.recon_inner_notebook.add(self.recon_subtab_insights, text="Insights")
        self.recon_inner_notebook.add(self.recon_subtab_notes, text="Credit/Debit Notes")

        self._build_details_subtab(self.recon_subtab_details)
        self._build_insights_subtab(self.recon_subtab_insights)
        self._build_notes_details_subtab(self.recon_subtab_notes)

        self._results = None
        self._note_results = None

    def _build_details_subtab(self, t):
        filt = ttk.Frame(t)
        filt.pack(fill="x", pady=(0, 8))
        ttk.Label(filt, text="View category:").pack(side="left")
        self.category_var = tk.StringVar(value="value_tax_mismatches")
        self.category_combo = ttk.Combobox(filt, textvariable=self.category_var, state="readonly", width=45, values=[
            "value_tax_mismatches", "probable_matches", "only_in_gstr2a",
            "only_in_purchase_register", "matched_clean", "multi_rate_reference",
        ])
        self.category_combo.pack(side="left", padx=8)
        self.category_combo.bind("<<ComboboxSelected>>", lambda e: self._render_category())

        result_container, self.result_tree = _make_scrollable_tree(t, (), height=16)
        result_container.pack(fill="both", expand=True, pady=(0, 10))

        btns = ttk.Frame(t)
        btns.pack(fill="x")
        ttk.Button(btns, text="Export Full Report to Excel...", command=self._export_report).pack(side="left")

    def _build_notes_details_subtab(self, t):
        info_lbl = ttk.Label(t, text="A Tally 'Credit Note' matches a supplier's GSTR 'Debit Note' filing, "
                              "and vice versa — the two sides label the same document oppositely, "
                              "so the GSTR and Books columns below will show opposite types on a match. "
                              "Included automatically in 'Export Full Report to Excel...' on the "
                              "Category Details tab (Notes_* sheets) — no separate export needed here.",
                              style="Muted.TLabel", justify="left")
        info_lbl.pack(anchor="w", pady=(0, 8))
        _bind_dynamic_wraplength(info_lbl, t)

        self.note_summary_var = tk.StringVar(value="Run Reconciliation above to see credit/debit note results.")
        note_summary_lbl = ttk.Label(t, textvariable=self.note_summary_var, style="Summary.TLabel", justify="left")
        note_summary_lbl.pack(anchor="w", pady=(0, 10))
        _bind_dynamic_wraplength(note_summary_lbl, t)

        filt = ttk.Frame(t)
        filt.pack(fill="x", pady=(0, 8))
        ttk.Label(filt, text="View category:").pack(side="left")
        self.note_category_var = tk.StringVar(value="value_tax_mismatches")
        self.note_category_combo = ttk.Combobox(filt, textvariable=self.note_category_var, state="readonly", width=45, values=[
            "value_tax_mismatches", "probable_matches", "only_in_gstr2b_cdnr",
            "only_in_note_register", "matched_clean",
        ])
        self.note_category_combo.pack(side="left", padx=8)
        self.note_category_combo.bind("<<ComboboxSelected>>", lambda e: self._render_note_category())

        note_result_container, self.note_result_tree = _make_scrollable_tree(t, (), height=16)
        note_result_container.pack(fill="both", expand=True, pady=(0, 10))

    def _render_note_category(self):
        self.note_result_tree.delete(*self.note_result_tree.get_children())
        self.note_result_tree["columns"] = ()
        if not self._note_results:
            return
        cat = self.note_category_var.get()
        df = self._note_results.get(cat)
        if df is None or len(df) == 0:
            self.note_result_tree["columns"] = ("info",)
            self.note_result_tree.heading("info", text="Result")
            self.note_result_tree.insert("", "end", values=("No rows in this category.",))
            return
        spec = _NOTE_CATEGORY_COLUMNS.get(cat)
        if spec:
            cols, rename = spec
            cols = [c for c in cols if c in df.columns]
            df = df[cols].rename(columns=rename)
        display_cols = list(df.columns)[:17]
        self.note_result_tree["columns"] = display_cols
        for c in display_cols:
            self.note_result_tree.heading(c, text=c)
            self.note_result_tree.column(c, width=_col_width(c), anchor="w")
        for _, row in df.head(500).iterrows():
            self.note_result_tree.insert("", "end", values=[row[c] for c in display_cols])

    def _refresh_recon_tab(self):
        if not self.current_client_id:
            return
        fys = service.list_available_fys(self.current_client_id)
        self.fy_combo.configure(values=fys)
        if fys:
            self.fy_combo.current(0)
        else:
            self.fy_var.set("")

        months = db.list_available_months(self.current_client_id)
        self.month_combo.configure(values=months)
        if months:
            self.month_combo.current(0)
        else:
            self.month_var.set("")

        self._results = None
        self._current_fy_meta = None
        self._note_results = None
        self._current_note_meta = None
        self.result_tree.delete(*self.result_tree.get_children())
        self.note_result_tree.delete(*self.note_result_tree.get_children())
        self.recon_summary_var.set("")
        self.recon_cutoff_var.set("")
        self.note_summary_var.set("Run Reconciliation above to see credit/debit note results.")
        self._reset_analysis_tab()

    def _run_reconciliation(self):
        if not self.fy_var.get():
            messagebox.showwarning("No financial year", "Upload at least one Purchase Register file first.")
            return
        try:
            outcome = service.run_fy_reconciliation(self.current_client_id, self.fy_var.get())
        except ValueError as e:
            messagebox.showwarning("Cannot reconcile", str(e))
            return
        except Exception as e:
            messagebox.showerror("Reconciliation failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return

        try:
            note_outcome = service.run_fy_note_reconciliation(self.current_client_id, self.fy_var.get())
        except Exception as e:
            note_outcome = None
            self._note_results = None
            self.note_result_tree.delete(*self.note_result_tree.get_children())
            self.note_summary_var.set(f"Credit/debit note reconciliation failed: {e}")

        if note_outcome is not None:
            self._note_results = note_outcome["results"]
            self._current_note_meta = note_outcome["meta"]
            if self._current_note_meta.get("has_data"):
                nr = self._note_results
                self.note_summary_var.set(
                    f"FY {self._current_note_meta['fy_label']}  |  Matched clean: {len(nr['matched_clean'])}   |   "
                    f"Mismatches: {len(nr['value_tax_mismatches'])}   |   Probable matches: {len(nr['probable_matches'])}   |   "
                    f"Only in GSTR-2B B2B-CDNR: {len(nr['only_in_gstr2b_cdnr'])}   |   "
                    f"Only in Note Register: {len(nr['only_in_note_register'])}   |   "
                    f"Pending conflicts (excluded): {self._current_note_meta['pending_conflicts_count']}"
                )
            else:
                self.note_summary_var.set(
                    f"No credit/debit note data uploaded for FY {self._current_note_meta['fy_label']} "
                    f"(period window {self._current_note_meta['period_window_display']})."
                )
            self._render_note_category()

        self._results = outcome["results"]
        self._current_fy_meta = outcome["meta"]
        r = self._results
        self.recon_summary_var.set(
            f"FY {self._current_fy_meta['fy_label']}  |  Matched clean: {len(r['matched_clean'])}   |   "
            f"Mismatches: {len(r['value_tax_mismatches'])}   |   Probable matches: {len(r['probable_matches'])}   |   "
            f"Only in GSTR-2A: {len(r['only_in_gstr2a'])}   |   Only in books: {len(r['only_in_purchase_register'])}   |   "
            f"Pending conflicts (excluded): {self._current_fy_meta['pending_conflicts_count']}"
        )
        days = self._current_fy_meta["days_until_cutoff"]
        snap_word = "snapshot" if self._current_fy_meta["gstr2a_snapshot_count"] == 1 else "snapshots"
        cutoff_text = (f"({days} days remaining)" if days >= 0 else f"({abs(days)} days PAST cutoff)")
        self.recon_cutoff_var.set(
            f"Late-filing cutoff: {self._current_fy_meta['cutoff_date']}  {cutoff_text}   |   "
            f"GSTR-2A data merged from {self._current_fy_meta['gstr2a_snapshot_count']} {snap_word} "
            f"(most recent uploaded {self._current_fy_meta['gstr2a_latest_uploaded_at']})"
        )
        self._render_category()
        self._render_analysis()

    def _render_category(self):
        self.result_tree.delete(*self.result_tree.get_children())
        self.result_tree["columns"] = ()
        if not self._results:
            return
        cat = self.category_var.get()
        df = self._results.get(cat)
        if df is None or len(df) == 0:
            self.result_tree["columns"] = ("info",)
            self.result_tree.heading("info", text="Result")
            self.result_tree.insert("", "end", values=("No rows in this category.",))
            return
        spec = _CATEGORY_COLUMNS.get(cat)
        if spec:
            cols, rename = spec
            cols = [c for c in cols if c in df.columns]
            df = df[cols].rename(columns=rename)
        display_cols = list(df.columns)[:17]
        self.result_tree["columns"] = display_cols
        for c in display_cols:
            self.result_tree.heading(c, text=c)
            self.result_tree.column(c, width=_col_width(c), anchor="w")
        for _, row in df.head(500).iterrows():
            self.result_tree.insert("", "end", values=[row[c] for c in display_cols])

    def _build_insights_subtab(self, t):
        self.analysis_info_var = tk.StringVar(value="Run a reconciliation above to see insights.")
        ttk.Label(t, textvariable=self.analysis_info_var, style="Muted.TLabel").pack(anchor="w", pady=(0, 14))

        kpi_row = ttk.Frame(t)
        kpi_row.pack(fill="x", pady=(0, 16))
        self.kpi_vars = {}
        kpi_defs = [
            ("purchase_total", "Total Purchase Value (FY)"),
            ("matched_value", "Matched Value"),
            ("itc_at_risk", "ITC at Risk (Only in Books)"),
            ("mismatch_value", "Mismatch Value (abs)"),
        ]
        for i, (key, label) in enumerate(kpi_defs):
            tile = ttk.Frame(kpi_row, relief="solid", borderwidth=1, padding=14)
            tile.pack(side="left", fill="both", expand=True, padx=(0 if i == 0 else 10, 0))
            var = tk.StringVar(value="—")
            ttk.Label(tile, textvariable=var, font=(UI_FONT, 18, "bold")).pack(anchor="w")
            ttk.Label(tile, text=label, style="Muted.TLabel").pack(anchor="w", pady=(2, 0))
            self.kpi_vars[key] = var

        # Charts are stacked full-width in a scrollable area rather than
        # squeezed side-by-side, so nothing gets crunched regardless of window
        # size — the supplier chart in particular needs real vertical room for
        # up to 8 bars with names and value labels.
        scroll_container, charts_area = _make_scrollable_frame(t)
        scroll_container.pack(fill="both", expand=True)

        self.fig_categories = Figure(figsize=(10.5, 3.2), dpi=96, facecolor=CHART_SURFACE)
        self.ax_categories = self.fig_categories.add_subplot(111)
        self.canvas_categories = FigureCanvasTkAgg(self.fig_categories, master=charts_area)
        self.canvas_categories.get_tk_widget().pack(fill="both", expand=True, pady=(0, 16))

        self.fig_monthly = Figure(figsize=(10.5, 3.0), dpi=96, facecolor=CHART_SURFACE)
        self.ax_monthly = self.fig_monthly.add_subplot(111)
        self.canvas_monthly = FigureCanvasTkAgg(self.fig_monthly, master=charts_area)
        self.canvas_monthly.get_tk_widget().pack(fill="both", expand=True, pady=(0, 16))

        self.fig_suppliers = Figure(figsize=(10.5, 4.6), dpi=96, facecolor=CHART_SURFACE)
        self.ax_suppliers = self.fig_suppliers.add_subplot(111)
        self.canvas_suppliers = FigureCanvasTkAgg(self.fig_suppliers, master=charts_area)
        self.canvas_suppliers.get_tk_widget().pack(fill="both", expand=True)

        self._reset_analysis_tab()

    def _reset_analysis_tab(self):
        self.analysis_info_var.set("Run a reconciliation above to see insights.")
        for var in self.kpi_vars.values():
            var.set("—")
        for ax, canvas in ((self.ax_categories, self.canvas_categories),
                           (self.ax_monthly, self.canvas_monthly),
                           (self.ax_suppliers, self.canvas_suppliers)):
            ax.clear()
            ax.axis("off")
            canvas.draw()

    def _style_axes(self, ax):
        _style_axes(ax)

    def _render_analysis(self):
        if not self._results:
            self._reset_analysis_tab()
            return
        r = self._results
        meta = self._current_fy_meta
        fy_label = meta["fy_label"]
        self.analysis_info_var.set(f"FY {fy_label}")

        matched_value = r["matched_clean"]["invoice_value"].sum() if len(r["matched_clean"]) else 0
        itc_at_risk = r["only_in_purchase_register"]["gross_total"].sum() if len(r["only_in_purchase_register"]) else 0
        mismatch_value = r["value_tax_mismatches"]["diff_value"].abs().sum() if len(r["value_tax_mismatches"]) else 0
        pr_fy_df = db.get_purchase_entries_by_fy(self.current_client_id, fy_label)
        purchase_total = pr_fy_df["gross_total"].sum() if len(pr_fy_df) else 0

        self.kpi_vars["purchase_total"].set(format_inr(purchase_total))
        self.kpi_vars["matched_value"].set(format_inr(matched_value))
        self.kpi_vars["itc_at_risk"].set(format_inr(itc_at_risk))
        self.kpi_vars["mismatch_value"].set(format_inr(mismatch_value))

        # Category counts: each bar is a distinct, named category -> categorical color.
        ax = self.ax_categories
        ax.clear()
        ax.axis("on")
        labels = ["Matched\nClean", "Value/Tax\nMismatches", "Probable\nMatches",
                  "Only in\nGSTR-2A", "Only in\nBooks", "Pending\nConflicts"]
        values = [len(r["matched_clean"]), len(r["value_tax_mismatches"]), len(r["probable_matches"]),
                  len(r["only_in_gstr2a"]), len(r["only_in_purchase_register"]), meta["pending_conflicts_count"]]
        bars = ax.bar(labels, values, color=CHART_CATEGORICAL, width=0.6)
        for bar, val in zip(bars, values):
            ax.annotate(str(val), (bar.get_x() + bar.get_width() / 2, bar.get_height()),
                        ha="center", va="bottom", fontsize=8, color=CHART_INK)
        ax.set_title("Reconciliation Categories", fontsize=10, fontweight="bold", loc="left")
        ax.grid(axis="y", color=CHART_GRID, linewidth=0.8)
        ax.set_axisbelow(True)
        self._style_axes(ax)
        self.fig_categories.tight_layout()
        self.canvas_categories.draw()

        # Monthly Purchase Register totals: a single series' magnitude -> one sequential hue.
        ax = self.ax_monthly
        ax.clear()
        ax.axis("on")
        if len(pr_fy_df):
            monthly = pr_fy_df.groupby("entry_month")["gross_total"].sum().sort_index()
            month_labels = [f"{m[5:7]}/{m[2:4]}" for m in monthly.index]
            ax.bar(month_labels, monthly.values, color=CHART_SEQUENTIAL, width=0.6)
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_inr_short(x)))
        else:
            ax.text(0.5, 0.5, "No Purchase Register data for this FY", ha="center", va="center",
                    color=CHART_MUTED, fontsize=9, transform=ax.transAxes)
        ax.set_title("Monthly Purchase Register Total", fontsize=10, fontweight="bold", loc="left")
        ax.grid(axis="y", color=CHART_GRID, linewidth=0.8)
        ax.set_axisbelow(True)
        self._style_axes(ax)
        self.fig_monthly.tight_layout()
        self.canvas_monthly.draw()

        # Top suppliers only in the Purchase Register (ITC at risk): a ranking of one
        # series -> one sequential hue, not a color per bar (color follows entity, not rank).
        ax = self.ax_suppliers
        ax.clear()
        ax.axis("on")
        only_pr = r["only_in_purchase_register"]
        if len(only_pr):
            top = only_pr.groupby("particulars")["gross_total"].sum().sort_values(ascending=False).head(8)
            top = top.sort_values(ascending=True)
            names = [n if len(n) <= 24 else n[:21] + "…" for n in top.index]
            ax.barh(names, top.values, color=CHART_SEQUENTIAL, height=0.5)
            ax.margins(x=0.16)
            for i, val in enumerate(top.values):
                ax.annotate(format_inr_short(val), (val, i), ha="left", va="center", fontsize=8,
                            color=CHART_INK, xytext=(6, 0), textcoords="offset points")
        else:
            ax.text(0.5, 0.5, "Nothing only in the Purchase Register for this FY", ha="center", va="center",
                    color=CHART_MUTED, fontsize=9, transform=ax.transAxes)
        ax.set_title("Top Suppliers Not Yet in GSTR-2A/2B (ITC at Risk)", fontsize=10, fontweight="bold", loc="left")
        ax.grid(axis="x", color=CHART_GRID, linewidth=0.8)
        ax.set_axisbelow(True)
        self._style_axes(ax)
        self.fig_suppliers.tight_layout()
        self.canvas_suppliers.draw()

    def _export_report(self):
        if not self._results:
            messagebox.showwarning("Run reconciliation first", "Click 'Run Reconciliation' before exporting.")
            return
        client = db.get_client(self.current_client_id)
        fy_label = self._current_fy_meta["fy_label"]
        default_name = f"{client['name'].replace(' ', '_')}_GST_Reconciliation_FY{fy_label}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save reconciliation report", initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            service.export_fy_reconciliation(self.current_client_id, fy_label, path)
        except Exception as e:
            messagebox.showerror("Export failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return
        self._refresh_reports_tab()
        if messagebox.askyesno("Report saved", f"Report saved to:\n{path}\n\nOpen it now?"):
            open_file(path)

    def _export_month(self):
        if not self.month_var.get():
            messagebox.showwarning("No month available", "Upload a Purchase Register file first.")
            return
        client = db.get_client(self.current_client_id)
        year_month = self.month_var.get()
        default_name = f"{client['name'].replace(' ', '_')}_Purchases_{year_month}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save month export", initialfile=default_name,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            service.export_purchase_month(self.current_client_id, year_month, path)
        except Exception as e:
            messagebox.showerror("Export failed", f"{e}\n\n{traceback.format_exc()[-800:]}")
            return
        if messagebox.askyesno("Export saved", f"Saved to:\n{path}\n\nOpen it now?"):
            open_file(path)

    # ---------------- Reports tab ----------------

    def _build_reports_tab(self):
        t = self.tab_reports
        ttk.Label(t, text="Past reconciliation reports for this client",
                  style="Section.TLabel").pack(anchor="w", pady=(0, 6))
        cols = ("Run Date", "FY", "GSTR-2A Snapshot", "Purchase Data As Of", "File")
        reports_container, self.reports_tree = _make_scrollable_tree(t, cols, height=16)
        reports_container.pack(fill="both", expand=True, pady=(0, 10))
        self.reports_tree.bind("<Double-1>", lambda e: self._view_report_insights())
        btns = ttk.Frame(t)
        btns.pack(fill="x")
        ttk.Button(btns, text="View Insights & Details...", command=self._view_report_insights).pack(side="left")
        ttk.Button(btns, text="Open Excel File", command=self._open_selected_report).pack(side="left", padx=8)

    def _refresh_reports_tab(self):
        if not self.current_client_id:
            return
        self.reports_tree.delete(*self.reports_tree.get_children())
        for r in db.list_recon_runs(self.current_client_id):
            self.reports_tree.insert("", "end", values=(
                r["run_at"], r["fy_label"] or "n/a", f"#{r['snapshot_id']}", r["purchase_asof"], r["report_path"],
            ))

    def _selected_report_path(self):
        sel = self.reports_tree.selection()
        if not sel:
            messagebox.showwarning("Nothing selected", "Select a report first.")
            return None
        path = self.reports_tree.item(sel[0])["values"][4]
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"This report file no longer exists at:\n{path}")
            return None
        return path

    def _open_selected_report(self):
        path = self._selected_report_path()
        if path:
            open_file(path)

    def _view_report_insights(self):
        path = self._selected_report_path()
        if path:
            ReportViewerWindow(self, path)


def main():
    _enable_dpi_awareness()
    if not licensing.is_activated():
        activation = ActivationScreen()
        activation.mainloop()
        if not activation.activated:
            return

    login = LoginScreen()
    login.mainloop()
    if not login.unlocked:
        return
    app = App()
    app.mainloop()
