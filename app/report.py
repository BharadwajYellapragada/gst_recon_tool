"""
Builds the multi-sheet, formatted Excel reconciliation report from the
dict of DataFrames produced by reconcile.run_reconciliation().
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color='1F4E78')
SUB_FONT = Font(name=FONT, size=10, italic=True, color='595959')
NORMAL = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
RED_FONT = Font(name=FONT, size=10, color='9C0006')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def _autosize(ws, df, max_width=40):
    for i, col in enumerate(df.columns, start=1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:500]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen + 2, 10), max_width)


def _write_df(ws, df, start_row=1, money_cols=None):
    money_cols = money_cols or []
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    _style_header(ws, start_row, len(df.columns))
    ws.row_dimensions[start_row].height = 30
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=row[col])
            cell.font = NORMAL
            cell.border = BORDER
            if col in money_cols:
                cell.number_format = '#,##0.00'
    _autosize(ws, df)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df)


def build_report(client_name, client_gstin, snapshot_info, purchase_asof, results, output_path, fy_meta=None,
                  note_results=None, note_meta=None):
    """
    client_name, client_gstin: str
    snapshot_info: dict with keys uploaded_at, generation_date, period_min, period_max, row_count
    purchase_asof: str (latest purchase entry date included)
    results: dict of DataFrames from reconcile.run_reconciliation()
    output_path: full path to save the .xlsx
    fy_meta: optional dict from service.run_fy_reconciliation()'s 'meta' — if given, the
             Summary sheet is scoped to that financial year and its late-filing window.
    note_results: optional dict of DataFrames from reconcile.run_note_reconciliation()
                  (via service.run_fy_note_reconciliation()) — if given (and
                  note_meta['has_data'] is True), adds a Credit/Debit Note section to
                  the Summary sheet plus dedicated Notes_* sheets.
    note_meta: optional dict from service.run_fy_note_reconciliation()'s 'meta'.
    """
    wb = Workbook()

    matched_clean = results["matched_clean"]
    mismatches = results["value_tax_mismatches"]
    probable = results["probable_matches"]
    only_g2a = results["only_in_gstr2a"]
    only_pr = results["only_in_purchase_register"]
    multi_rate = results["multi_rate_reference"]
    true_dups = results["true_duplicates_gstr2a"]

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = "GST Reconciliation Report"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"{client_name}  |  GSTIN {client_gstin}"
    ws["B3"].font = SUB_FONT
    if fy_meta:
        ws["B4"] = (f"Financial Year {fy_meta['fy_label']}  \u2014  Purchase Register for this FY cross-checked "
                    f"against GSTR-2A periods {fy_meta['period_window_display']} "
                    f"({fy_meta['late_filing_months']}-month late-filing allowance)")
        ws["B4"].font = SUB_FONT
        snap_word = "snapshot" if fy_meta["gstr2a_snapshot_count"] == 1 else "snapshots"
        ws["B5"] = (f"GSTR-2A data merged from {fy_meta['gstr2a_snapshot_count']} portal {snap_word} "
                    f"(most recent uploaded {fy_meta['gstr2a_latest_uploaded_at']}, "
                    f"portal generation date: {fy_meta['gstr2a_latest_generation_date'] or 'n/a'})")
        ws["B5"].font = SUB_FONT
        cutoff_note = (f"Late-filing cutoff for this FY: {fy_meta['cutoff_date']}"
                       + (f"  \u2014  {fy_meta['days_until_cutoff']} days remaining" if fy_meta['days_until_cutoff'] >= 0
                          else f"  \u2014  {abs(fy_meta['days_until_cutoff'])} days past cutoff"))
        ws["B6"] = cutoff_note
        ws["B6"].font = Font(name=FONT, size=10, bold=True, color=('9C0006' if fy_meta['days_until_cutoff'] < 0 else '006100'))
        header_row_offset = 2
        if fy_meta.get("pending_conflicts_count"):
            ws["B7"] = (f"\u26a0 {fy_meta['pending_conflicts_count']} uploaded Purchase Register row(s) are "
                        f"awaiting an Overwrite/Ignore decision and are EXCLUDED from the figures below "
                        f"until resolved. See Pending_Conflicts sheet.")
            ws["B7"].font = Font(name=FONT, size=10, bold=True, color='9C0006')
            header_row_offset = 3
    else:
        ws["B4"] = (f"GSTR-2A snapshot uploaded {snapshot_info.get('uploaded_at','')} "
                    f"(portal generation date: {snapshot_info.get('generation_date','') or 'n/a'}), "
                    f"periods {snapshot_info.get('period_min','')}\u2013{snapshot_info.get('period_max','')}")
        ws["B4"].font = SUB_FONT
        ws["B5"] = f"Purchase Register data as of latest entry: {purchase_asof or 'n/a'}"
        ws["B5"].font = SUB_FONT
        header_row_offset = 0

    recon_rows = [
        ("Category", "Count", "Value Impact (Rs)", "Sheet"),
        ("Matched & fully agree (within Rs 2 tolerance)", len(matched_clean), "", "Matched_Clean"),
        ("Matched but value/tax mismatch", len(mismatches),
         round(mismatches["diff_value"].abs().sum(), 2) if len(mismatches) else 0, "Value_Tax_Mismatches"),
        ("Probable match \u2014 invoice no. differs, GSTIN+date+amount agree", len(probable), "", "Probable_Matches"),
        ("In GSTR-2A only \u2014 not found in books", len(only_g2a),
         round(only_g2a["invoice_value"].sum(), 2) if len(only_g2a) else 0, "Only_In_GSTR2A"),
        ("In Purchase Register only \u2014 supplier hasn't filed (ITC at risk)", len(only_pr),
         round(only_pr["gross_total"].sum(), 2) if len(only_pr) else 0, "Only_In_PurchaseRegister"),
        ("True duplicate rows in GSTR-2A", len(true_dups), "", "Duplicates_GSTR2A"),
        ("Multi-rate invoices in GSTR-2A (legitimate line-splits, FYI)", len(multi_rate["match_key"].unique()) if len(multi_rate) else 0, "", "MultiRate_Reference"),
    ]
    r0 = 7 + header_row_offset
    for i, rrow in enumerate(recon_rows):
        for j, val in enumerate(rrow):
            cell = ws.cell(row=r0 + i, column=2 + j, value=val)
            if i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = NORMAL
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
            cell.border = BORDER
    for col, w in zip("BCDE", [65, 10, 20, 28]):
        ws.column_dimensions[col].width = w

    r_next = r0 + len(recon_rows) + 1
    if note_meta and note_meta.get("has_data") and note_results is not None:
        n_matched = note_results["matched_clean"]
        n_mismatches = note_results["value_tax_mismatches"]
        n_probable = note_results["probable_matches"]
        n_only_gstr = note_results["only_in_gstr2b_cdnr"]
        n_only_books = note_results["only_in_note_register"]

        ws[f"B{r_next}"] = "Credit / Debit Note Reconciliation"
        ws[f"B{r_next}"].font = Font(name=FONT, bold=True, size=12, color='1F4E78')
        ws[f"B{r_next + 1}"] = (
            "A Tally 'Credit Note' register entry matches a supplier's GSTR-2B 'Debit Note' filing, "
            "and vice versa — the two sides label the same document oppositely (see the GSTR/Books "
            "columns on each Notes_* sheet)."
        )
        ws[f"B{r_next + 1}"].font = SUB_FONT
        note_rows = [
            ("Category", "Count", "Value Impact (Rs)", "Sheet"),
            ("Matched & fully agree (within Rs 2 tolerance)", len(n_matched), "", "Notes_Matched_Clean"),
            ("Matched but value/tax mismatch", len(n_mismatches),
             round(n_mismatches["diff_value"].abs().sum(), 2) if len(n_mismatches) else 0, "Notes_Value_Tax_Mismatches"),
            ("Probable match — note no. differs, GSTIN+date+amount agree", len(n_probable), "", "Notes_Probable_Matches"),
            ("In GSTR-2B B2B-CDNR only — not found in books", len(n_only_gstr),
             round(n_only_gstr["note_value"].sum(), 2) if len(n_only_gstr) else 0, "Notes_Only_In_GSTR2B_CDNR"),
            ("In Note Register only — supplier hasn't filed", len(n_only_books),
             round(n_only_books["gross_total"].sum(), 2) if len(n_only_books) else 0, "Notes_Only_In_NoteRegister"),
        ]
        nr0 = r_next + 3
        for i, rrow in enumerate(note_rows):
            for j, val in enumerate(rrow):
                cell = ws.cell(row=nr0 + i, column=2 + j, value=val)
                if i == 0:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                else:
                    cell.font = NORMAL
                    if isinstance(val, float):
                        cell.number_format = "#,##0.00"
                cell.border = BORDER
        note_pending = note_results.get("pending_conflicts")
        if note_pending:
            pr_row = nr0 + len(note_rows) + 1
            ws[f"B{pr_row}"] = (f"⚠ {len(note_pending)} uploaded Note Register row(s) are awaiting an "
                                 f"Overwrite/Ignore decision and are EXCLUDED from the figures above until "
                                 f"resolved. See Notes_Pending_Conflicts sheet.")
            ws[f"B{pr_row}"].font = Font(name=FONT, size=10, bold=True, color='9C0006')

    # ---------------- Value_Tax_Mismatches ----------------
    ws2 = wb.create_sheet("Value_Tax_Mismatches")
    if len(mismatches):
        cols = ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_no_pr", "invoice_date_2a",
                "invoice_value", "gross_total", "diff_value", "igst_2a", "igst_pr", "diff_igst",
                "cgst_2a", "cgst_pr", "diff_cgst", "sgst_2a", "sgst_pr", "diff_sgst"]
        df = mismatches[cols].rename(columns={
            "gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No (2A)",
            "invoice_no_pr": "Invoice No (Books)", "invoice_date_2a": "Invoice Date",
            "invoice_value": "Invoice Value (2A)", "gross_total": "Gross Total (Books)",
            "diff_value": "Diff_Value", "igst_2a": "IGST (2A)", "igst_pr": "IGST (Books)",
            "diff_igst": "Diff_IGST", "cgst_2a": "CGST (2A)", "cgst_pr": "CGST (Books)",
            "diff_cgst": "Diff_CGST", "sgst_2a": "SGST (2A)", "sgst_pr": "SGST (Books)", "diff_sgst": "Diff_SGST",
        }).sort_values("Diff_Value", key=abs, ascending=False)
        money_cols = ["Invoice Value (2A)", "Gross Total (Books)", "Diff_Value", "IGST (2A)", "IGST (Books)",
                      "Diff_IGST", "CGST (2A)", "CGST (Books)", "Diff_CGST", "SGST (2A)", "SGST (Books)", "Diff_SGST"]
        lastrow = _write_df(ws2, df, money_cols=money_cols)
        diff_idx = [df.columns.get_loc(c) + 1 for c in ["Diff_Value", "Diff_IGST", "Diff_CGST", "Diff_SGST"]]
        for r in range(2, lastrow + 1):
            for c in diff_idx:
                cell = ws2.cell(row=r, column=c)
                if cell.value and abs(cell.value) > 2:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
    else:
        ws2["A1"] = "No value/tax mismatches found among matched invoices."
        ws2["A1"].font = BOLD

    # ---------------- Probable_Matches ----------------
    ws3 = wb.create_sheet("Probable_Matches")
    ws3["A1"] = ("Same GSTIN, same invoice date, same amount (\u00b1Re.1) but invoice numbers differ. "
                 "Likely a typo/format difference on one side \u2014 verify manually.")
    ws3["A1"].font = SUB_FONT
    if len(probable):
        cols = ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_no_pr", "invoice_date_2a",
                "invoice_value", "gross_total"]
        df = probable[cols].rename(columns={
            "gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No (2A)",
            "invoice_no_pr": "Invoice No (Books)", "invoice_date_2a": "Invoice Date",
            "invoice_value": "Invoice Value (2A)", "gross_total": "Gross Total (Books)",
        })
        lastrow = _write_df(ws3, df, start_row=3, money_cols=["Invoice Value (2A)", "Gross Total (Books)"])
        for r in range(4, lastrow + 2):
            ws3.cell(row=r, column=3).fill = YELLOW_FILL
            ws3.cell(row=r, column=4).fill = YELLOW_FILL

    # ---------------- Only_In_GSTR2A ----------------
    ws4 = wb.create_sheet("Only_In_GSTR2A")
    if len(only_g2a):
        cols = ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_date_2a", "invoice_value",
                "taxable_value", "igst_2a", "cgst_2a", "sgst_2a", "itc_available", "period"]
        df = only_g2a[cols].rename(columns={
            "gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No",
            "invoice_date_2a": "Invoice Date", "invoice_value": "Invoice Value",
            "taxable_value": "Taxable Value", "igst_2a": "IGST", "cgst_2a": "CGST", "sgst_2a": "SGST",
            "itc_available": "ITC Available",
        }).sort_values("Invoice Value", ascending=False)
        money_cols = ["Invoice Value", "Taxable Value", "IGST", "CGST", "SGST"]
        lastrow = _write_df(ws4, df, money_cols=money_cols)
        for r in range(2, lastrow + 1):
            ws4.cell(row=r, column=1).fill = YELLOW_FILL
    else:
        ws4["A1"] = "Nothing found only in GSTR-2A."
        ws4["A1"].font = BOLD

    # ---------------- Only_In_PurchaseRegister ----------------
    ws5 = wb.create_sheet("Only_In_PurchaseRegister")
    if len(only_pr):
        has_flag = "past_6_month_window" in only_pr.columns
        cols = ["gstin_pr", "particulars", "invoice_no_pr", "invoice_date_pr", "gross_total",
                "igst_pr", "cgst_pr", "sgst_pr"]
        rename_map = {
            "gstin_pr": "GSTIN", "particulars": "Supplier Name", "invoice_no_pr": "Invoice No",
            "invoice_date_pr": "Invoice Date", "gross_total": "Gross Total", "igst_pr": "IGST",
            "cgst_pr": "CGST", "sgst_pr": "SGST",
        }
        if has_flag:
            cols.append("past_6_month_window")
            rename_map["past_6_month_window"] = "Past Late-Filing Cutoff?"
        df = only_pr[cols].rename(columns=rename_map).sort_values("Gross Total", ascending=False)
        money_cols = ["Gross Total", "IGST", "CGST", "SGST"]
        lastrow = _write_df(ws5, df, money_cols=money_cols)
        flag_col = df.columns.get_loc("Past Late-Filing Cutoff?") + 1 if has_flag else None
        for r in range(2, lastrow + 1):
            ws5.cell(row=r, column=1).fill = YELLOW_FILL
            if flag_col:
                cell = ws5.cell(row=r, column=flag_col)
                if cell.value:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
        if has_flag:
            ws5.insert_rows(1)
            ws5["A1"] = ("'Past Late-Filing Cutoff' = TRUE means the supplier's window to file this "
                         "invoice in their GSTR-1 (this FY + late-filing allowance) has closed \u2014 "
                         "highest priority for supplier follow-up.")
            ws5["A1"].font = SUB_FONT
    else:
        ws5["A1"] = "Nothing found only in the Purchase Register."
        ws5["A1"].font = BOLD

    # ---------------- Duplicates_GSTR2A ----------------
    ws6 = wb.create_sheet("Duplicates_GSTR2A")
    if len(true_dups):
        cols = ["period", "gstin", "supplier_name", "invoice_no", "invoice_date", "invoice_value",
                "rate", "taxable_value", "igst", "cgst", "sgst"]
        df = true_dups[cols].rename(columns={"supplier_name": "Supplier Name"})
        _write_df(ws6, df, money_cols=["invoice_value", "taxable_value", "igst", "cgst", "sgst"])
    else:
        ws6["A1"] = "No true duplicate rows found in this GSTR-2A snapshot."
        ws6["A1"].font = BOLD
        ws6["A2"] = ("Invoices appearing on multiple rows are multi-rate line splits from the portal "
                     "(normal, not an error) \u2014 see MultiRate_Reference.")
        ws6["A2"].font = SUB_FONT
        ws6.column_dimensions["A"].width = 100

    # ---------------- MultiRate_Reference ----------------
    ws7 = wb.create_sheet("MultiRate_Reference")
    if len(multi_rate):
        cols = ["period", "gstin", "supplier_name", "invoice_no", "invoice_date", "invoice_value",
                "rate", "taxable_value", "igst", "cgst", "sgst"]
        df = multi_rate[cols].rename(columns={"supplier_name": "Supplier Name"}).sort_values(["gstin", "invoice_no"])
        _write_df(ws7, df, money_cols=["invoice_value", "taxable_value", "igst", "cgst", "sgst"])

    # ---------------- Matched_Clean ----------------
    ws8 = wb.create_sheet("Matched_Clean")
    if len(matched_clean):
        cols = ["gstin_2a", "supplier_name", "invoice_no_2a", "invoice_date_2a", "invoice_value",
                "igst_2a", "cgst_2a", "sgst_2a", "itc_available"]
        df = matched_clean[cols].rename(columns={
            "gstin_2a": "GSTIN", "supplier_name": "Supplier Name", "invoice_no_2a": "Invoice No",
            "invoice_date_2a": "Invoice Date", "invoice_value": "Invoice Value", "igst_2a": "IGST",
            "cgst_2a": "CGST", "sgst_2a": "SGST", "itc_available": "ITC Available",
        })
        _write_df(ws8, df, money_cols=["Invoice Value", "IGST", "CGST", "SGST"])

    # ---------------- Pending_Conflicts ----------------
    pending = results.get("pending_conflicts")
    if pending:
        wsP = wb.create_sheet("Pending_Conflicts")
        wsP["A1"] = ("These invoices were re-uploaded with a DIFFERENT amount than what's already "
                     "stored. They are excluded from every figure in this report until you resolve "
                     "each one as Overwrite (use the new value) or Ignore (keep the stored value).")
        wsP["A1"].font = SUB_FONT
        cols = ["GSTIN", "Particulars", "Invoice No", "Stored: Gross Total", "New Upload: Gross Total",
                "Stored: CGST", "New Upload: CGST", "Stored: SGST", "New Upload: SGST",
                "Stored: IGST", "New Upload: IGST", "Pending Entry ID (for resolving)"]
        for j, c in enumerate(cols, start=1):
            wsP.cell(row=3, column=j, value=c)
        _style_header(wsP, 3, len(cols))
        r = 4
        for item in pending:
            p, s = item["pending"], item["stored"]
            vals = [p["gstin"], p["particulars"], p["invoice_no"],
                    s["gross_total"] if s else "", p["gross_total"],
                    s["cgst"] if s else "", p["cgst"],
                    s["sgst"] if s else "", p["sgst"],
                    s["igst"] if s else "", p["igst"],
                    p["id"]]
            for j, v in enumerate(vals, start=1):
                cell = wsP.cell(row=r, column=j, value=v)
                cell.font = NORMAL
                cell.border = BORDER
                cell.fill = YELLOW_FILL
            r += 1
        for i in range(1, len(cols) + 1):
            wsP.column_dimensions[get_column_letter(i)].width = 20

    order = ["Summary", "Value_Tax_Mismatches", "Probable_Matches", "Only_In_GSTR2A",
             "Only_In_PurchaseRegister", "Duplicates_GSTR2A", "MultiRate_Reference", "Matched_Clean"]
    if pending:
        order.insert(1, "Pending_Conflicts")

    if note_meta and note_meta.get("has_data") and note_results is not None:
        note_order = _write_note_sheets(wb, note_results)
        order += note_order

    wb._sheets = [wb[s] for s in order]
    wb.save(output_path)
    return output_path


def _write_note_sheets(wb, note_results):
    """Adds the Notes_* sheets (mirrors the invoice-side sheets above) for a
    credit/debit note reconciliation. Returns the sheet names in display order."""
    matched = note_results["matched_clean"]
    mismatches = note_results["value_tax_mismatches"]
    probable = note_results["probable_matches"]
    only_gstr = note_results["only_in_gstr2b_cdnr"]
    only_books = note_results["only_in_note_register"]
    pending = note_results.get("pending_conflicts")

    order = []

    ws = wb.create_sheet("Notes_Value_Tax_Mismatches")
    order.append("Notes_Value_Tax_Mismatches")
    if len(mismatches):
        cols = ["gstin_gstr", "supplier_name", "note_no", "voucher_no", "note_type", "book_note_type",
                "note_date", "note_value", "gross_total", "diff_value", "igst_gstr", "igst_books", "diff_igst",
                "cgst_gstr", "cgst_books", "diff_cgst", "sgst_gstr", "sgst_books", "diff_sgst"]
        df = mismatches[cols].rename(columns={
            "gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No (GSTR)",
            "voucher_no": "Voucher No (Books)", "note_type": "Note Type (GSTR)", "book_note_type": "Note Type (Books)",
            "note_date": "Note Date", "note_value": "Note Value (GSTR)", "gross_total": "Gross Total (Books)",
            "diff_value": "Diff_Value", "igst_gstr": "IGST (GSTR)", "igst_books": "IGST (Books)",
            "diff_igst": "Diff_IGST", "cgst_gstr": "CGST (GSTR)", "cgst_books": "CGST (Books)",
            "diff_cgst": "Diff_CGST", "sgst_gstr": "SGST (GSTR)", "sgst_books": "SGST (Books)", "diff_sgst": "Diff_SGST",
        }).sort_values("Diff_Value", key=abs, ascending=False)
        money_cols = ["Note Value (GSTR)", "Gross Total (Books)", "Diff_Value", "IGST (GSTR)", "IGST (Books)",
                      "Diff_IGST", "CGST (GSTR)", "CGST (Books)", "Diff_CGST", "SGST (GSTR)", "SGST (Books)", "Diff_SGST"]
        lastrow = _write_df(ws, df, money_cols=money_cols)
        diff_idx = [df.columns.get_loc(c) + 1 for c in ["Diff_Value", "Diff_IGST", "Diff_CGST", "Diff_SGST"]]
        for r in range(2, lastrow + 1):
            for c in diff_idx:
                cell = ws.cell(row=r, column=c)
                if cell.value and abs(cell.value) > 2:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
    else:
        ws["A1"] = "No value/tax mismatches found among matched credit/debit notes."
        ws["A1"].font = BOLD

    ws = wb.create_sheet("Notes_Probable_Matches")
    order.append("Notes_Probable_Matches")
    ws["A1"] = ("Same GSTIN, same date, same amount (matching the Tally↔GSTR note-type flip) but note "
                "numbers differ. Likely a typo/format difference on one side — verify manually.")
    ws["A1"].font = SUB_FONT
    if len(probable):
        cols = ["gstin_gstr", "supplier_name", "note_no", "voucher_no", "note_type", "book_note_type",
                "note_date", "note_value", "gross_total"]
        df = probable[cols].rename(columns={
            "gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No (GSTR)",
            "voucher_no": "Voucher No (Books)", "note_type": "Note Type (GSTR)", "book_note_type": "Note Type (Books)",
            "note_date": "Note Date", "note_value": "Note Value (GSTR)", "gross_total": "Gross Total (Books)",
        })
        lastrow = _write_df(ws, df, start_row=3, money_cols=["Note Value (GSTR)", "Gross Total (Books)"])
        for r in range(4, lastrow + 2):
            ws.cell(row=r, column=3).fill = YELLOW_FILL
            ws.cell(row=r, column=4).fill = YELLOW_FILL

    ws = wb.create_sheet("Notes_Only_In_GSTR2B_CDNR")
    order.append("Notes_Only_In_GSTR2B_CDNR")
    if len(only_gstr):
        cols = ["gstin_gstr", "supplier_name", "note_no", "note_type", "note_date", "note_value",
                "taxable_value", "igst_gstr", "cgst_gstr", "sgst_gstr", "itc_available", "period"]
        df = only_gstr[cols].rename(columns={
            "gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No",
            "note_type": "Note Type", "note_date": "Note Date", "note_value": "Note Value",
            "taxable_value": "Taxable Value", "igst_gstr": "IGST", "cgst_gstr": "CGST", "sgst_gstr": "SGST",
            "itc_available": "ITC Available",
        }).sort_values("Note Value", ascending=False)
        money_cols = ["Note Value", "Taxable Value", "IGST", "CGST", "SGST"]
        lastrow = _write_df(ws, df, money_cols=money_cols)
        for r in range(2, lastrow + 1):
            ws.cell(row=r, column=1).fill = YELLOW_FILL
    else:
        ws["A1"] = "Nothing found only in the GSTR-2B B2B-CDNR sheet."
        ws["A1"].font = BOLD

    ws = wb.create_sheet("Notes_Only_In_NoteRegister")
    order.append("Notes_Only_In_NoteRegister")
    if len(only_books):
        cols = ["gstin_books", "particulars", "voucher_no", "book_note_type", "voucher_date",
                "gross_total", "igst_books", "cgst_books", "sgst_books"]
        df = only_books[cols].rename(columns={
            "gstin_books": "GSTIN", "particulars": "Supplier Name", "voucher_no": "Voucher No",
            "book_note_type": "Note Type (Books)", "voucher_date": "Voucher Date", "gross_total": "Gross Total",
            "igst_books": "IGST", "cgst_books": "CGST", "sgst_books": "SGST",
        }).sort_values("Gross Total", ascending=False)
        money_cols = ["Gross Total", "IGST", "CGST", "SGST"]
        lastrow = _write_df(ws, df, money_cols=money_cols)
        for r in range(2, lastrow + 1):
            ws.cell(row=r, column=1).fill = YELLOW_FILL
    else:
        ws["A1"] = "Nothing found only in the Note Register."
        ws["A1"].font = BOLD

    ws = wb.create_sheet("Notes_Matched_Clean")
    order.append("Notes_Matched_Clean")
    if len(matched):
        cols = ["gstin_gstr", "supplier_name", "note_no", "note_type", "note_date", "note_value", "itc_available"]
        df = matched[cols].rename(columns={
            "gstin_gstr": "GSTIN", "supplier_name": "Supplier Name", "note_no": "Note No",
            "note_type": "Note Type", "note_date": "Note Date", "note_value": "Note Value",
            "itc_available": "ITC Available",
        })
        _write_df(ws, df, money_cols=["Note Value"])

    if pending:
        wsP = wb.create_sheet("Notes_Pending_Conflicts")
        order.insert(0, "Notes_Pending_Conflicts")
        wsP["A1"] = ("These credit/debit notes were re-uploaded with a DIFFERENT amount than what's "
                     "already stored. They are excluded from every figure in this report until you "
                     "resolve each one as Overwrite (use the new value) or Ignore (keep the stored value).")
        wsP["A1"].font = SUB_FONT
        cols = ["GSTIN", "Particulars", "Voucher No", "Stored: Gross Total", "New Upload: Gross Total",
                "Stored: CGST", "New Upload: CGST", "Stored: SGST", "New Upload: SGST",
                "Stored: IGST", "New Upload: IGST", "Pending Entry ID (for resolving)"]
        for j, c in enumerate(cols, start=1):
            wsP.cell(row=3, column=j, value=c)
        _style_header(wsP, 3, len(cols))
        r = 4
        for item in pending:
            p, s = item["pending"], item["stored"]
            vals = [p["gstin"], p["particulars"], p["voucher_no"],
                    s["gross_total"] if s else "", p["gross_total"],
                    s["cgst"] if s else "", p["cgst"],
                    s["sgst"] if s else "", p["sgst"],
                    s["igst"] if s else "", p["igst"],
                    p["id"]]
            for j, v in enumerate(vals, start=1):
                cell = wsP.cell(row=r, column=j, value=v)
                cell.font = NORMAL
                cell.border = BORDER
                cell.fill = YELLOW_FILL
            r += 1
        for i in range(1, len(cols) + 1):
            wsP.column_dimensions[get_column_letter(i)].width = 20

    return order
