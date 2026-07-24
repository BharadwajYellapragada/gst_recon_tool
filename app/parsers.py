"""
Parsers for the source file types:
  - GSTR-2A/2B export from the GST portal (sheet 'B2B', plus 'B2B-CDNR' for
    credit/debit notes when present)
  - Internal Purchase Register export (e.g. from Tally)
  - Internal Credit Note / Debit Note Register export (e.g. from Tally)

Both .xls (legacy, via xlrd) and .xlsx (via openpyxl) are supported.
"""
import re
import datetime
import pandas as pd

from . import fy_utils

G2A_COLS = ['Period', 'GSTIN', 'SupplierName', 'InvoiceNo', 'InvoiceType', 'InvoiceDate',
            'InvoiceValue', 'PlaceOfSupply', 'RCM', 'Rate', 'TaxableValue', 'IGST', 'CGST',
            'SGST', 'Cess', 'FilingDate', 'ITCAvailable', 'Reason']

CDNR_COLS = ['Period', 'GSTIN', 'SupplierName', 'NoteNo', 'NoteType', 'NoteDate', 'NoteValue',
             'PlaceOfSupply', 'RCM', 'TaxableValue', 'IGST', 'CGST', 'SGST', 'Cess',
             'FilingDate', 'ITCAvailable', 'Reason']

# GSTR-2B's B2B-CDNR sheet classifies a note from the SUPPLIER's filing side.
# A Tally buyer's-books "Credit Note" register entry is the mirror image of a
# supplier's GSTR "Debit Note" filing, and vice versa -- so matching a Tally
# note against its GSTR counterpart requires flipping the type, not comparing
# it directly.
TALLY_TO_GSTR_NOTE_TYPE = {"credit": "Debit Note", "debit": "Credit Note"}


class ParseError(Exception):
    pass


def _norm_inv(x):
    x = str(x).upper().strip()
    return re.sub(r'[^A-Z0-9]', '', x)


_PERIOD_RE = re.compile(r"([A-Za-z]{3,})[^0-9]*?(\d{2,4})")


def _normalize_period(s):
    """Canonicalizes the B2B sheet's period value to 'MMM-YY' (matching
    fy_utils.period_str), since GSTR-2A spells it 'APR-22' but GSTR-2B spells
    the same thing "Apr'25" — without this, FY-window filtering would silently
    exclude every GSTR-2B row from reconciliation."""
    s = str(s or "").strip()
    m = _PERIOD_RE.search(s)
    if not m:
        return s
    return f"{m.group(1)[:3].upper()}-{m.group(2)[-2:]}"


def _num(x):
    return x if isinstance(x, (int, float)) else 0


def _normalize_header(s):
    """Strips whitespace/newlines/currency symbols/punctuation so headers like
    'Taxable value (₹)' and 'Taxable Value (₹)' both become 'taxablevalue'."""
    return re.sub(r'[^a-z0-9]', '', str(s or "").lower())


def _gst_bucket(header_text):
    """Classifies a Tally register column as 'cgst'/'sgst'/'igst'/None by content,
    not just a fixed prefix. The same export can mix short names ('CGST@9%',
    'SGST @ 14%') with Tally's expanded ledger names ('IN: Central GST OP 2.50 %',
    'IN: State GST OP 2.50 %', 'IN: Intgrted GST OP 18.00 %' -- note the 'Intgrted'
    typo Tally itself ships with) depending on how the client's ledgers were set
    up, so a plain .startswith('cgst') misses the expanded-name columns entirely
    and silently drops that slab's tax from the total."""
    norm = _normalize_header(header_text)
    if not norm or "gst" not in norm:
        return None
    if norm.startswith("cgst") or "central" in norm:
        return "cgst"
    if norm.startswith("sgst") or "state" in norm:
        return "sgst"
    if norm.startswith("igst") or "integ" in norm or "intgr" in norm:
        return "igst"
    return None


def _combine_header_rows(row1, row2):
    """The B2B sheet's header spans two rows: a group header (e.g. 'Invoice details')
    whose sub-columns are named on the row below (e.g. 'Invoice number'). The
    sub-header is more specific, so it wins wherever present."""
    n = max(len(row1), len(row2))
    out = []
    for i in range(n):
        v2 = row2[i] if i < len(row2) else None
        v1 = row1[i] if i < len(row1) else None
        out.append(v2 if v2 not in (None, "") else v1)
    return out


def _open_book(path):
    if path.lower().endswith(".xls"):
        import xlrd
        return ("xlrd", xlrd.open_workbook(path))
    else:
        from openpyxl import load_workbook
        return ("openpyxl", load_workbook(path, read_only=True, data_only=True))


def parse_gstr2a(path):
    """
    Returns (invoices_df, generation_date_str).
    invoices_df has columns G2A_COLS + MatchKey, GSTIN uppercased.
    Raises ParseError with a user-facing message on bad file.

    Handles both GSTR-2A exports (Table 8A style, 'Period' as the first B2B
    column) and GSTR-2B exports (no leading Period column; several extra
    columns like IRN/Source) by locating each field by its header text
    instead of a fixed column position — the two reports don't share a
    column layout even though both use a 'B2B' sheet name.
    """
    kind, wb = _open_book(path)
    sheet_names = wb.sheet_names() if kind == "xlrd" else wb.sheetnames
    if "B2B" not in sheet_names:
        raise ParseError(
            "This doesn't look like a GSTR-2A/2B export — no 'B2B' sheet found. "
            "Please upload the .xls/.xlsx file downloaded from the GST portal's "
            "'Download GSTR-2A' or 'Download GSTR-2B' option."
        )

    generation_date = ""
    if "Read me" in sheet_names:
        try:
            if kind == "xlrd":
                rm = wb.sheet_by_name("Read me")
                rm_rows = [rm.row_values(r) for r in range(min(15, rm.nrows))]
            else:
                rm = wb["Read me"]
                rm_rows = [list(r) for r in rm.iter_rows(min_row=1, max_row=15, values_only=True)]
            for row in rm_rows:
                for i, cell in enumerate(row):
                    if cell and "generation" in str(cell).lower():
                        # The value sits 1 or 2 cells to the right depending on
                        # the report (GSTR-2A vs 2B lay out this sheet differently).
                        for offset in (1, 2):
                            if i + offset < len(row) and row[i + offset]:
                                generation_date = str(row[i + offset])
                                break
                    if generation_date:
                        break
                if generation_date:
                    break
        except Exception:
            pass

    if kind == "xlrd":
        ws = wb.sheet_by_name("B2B")
        all_rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        ws = wb["B2B"]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Locate the two-row header by its anchor column ('GSTIN of supplier') rather
    # than assuming a fixed row number.
    header_row_idx = None
    for i, row in enumerate(all_rows[:20]):
        cells = [_normalize_header(c) for c in row if c is not None]
        if any(c == "gstinofsupplier" for c in cells):
            header_row_idx = i
            break
    if header_row_idx is None or header_row_idx + 1 >= len(all_rows):
        raise ParseError(
            "Could not find the header row (expected a 'GSTIN of supplier' column) in the "
            "B2B sheet. This file's layout doesn't match a recognised GSTR-2A/2B export."
        )

    header = _combine_header_rows(all_rows[header_row_idx], all_rows[header_row_idx + 1])
    norm_header = [_normalize_header(h) for h in header]

    def find_col(*keywords):
        for i, h in enumerate(norm_header):
            if all(kw in h for kw in keywords):
                return i
        return None

    col_map = {
        "Period": find_col("period"),
        "GSTIN": find_col("gstin"),
        "SupplierName": find_col("legal"),
        "InvoiceNo": find_col("invoice", "number"),
        "InvoiceType": find_col("invoice", "type"),
        "InvoiceDate": find_col("invoice", "date"),
        "InvoiceValue": find_col("invoice", "value"),
        "PlaceOfSupply": find_col("place", "supply"),
        "RCM": find_col("reverse", "charge"),
        "Rate": find_col("rate"),
        "TaxableValue": find_col("taxable", "value"),
        "IGST": find_col("integrated", "tax"),
        "CGST": find_col("central", "tax"),
        "SGST": find_col("state", "tax"),
        "Cess": find_col("cess"),
        "FilingDate": find_col("filing", "date"),
        "ITCAvailable": find_col("itc"),
        "Reason": find_col("reason"),
    }
    required = ["Period", "GSTIN", "InvoiceNo", "InvoiceDate", "InvoiceValue",
                "TaxableValue", "IGST", "CGST", "SGST"]
    missing = [k for k in required if col_map[k] is None]
    if missing:
        raise ParseError(
            f"Could not find expected column(s) {', '.join(missing)} in the B2B sheet. "
            "This file's layout doesn't match a recognised GSTR-2A/2B export."
        )

    data_rows = all_rows[header_row_idx + 2:]
    rows = []
    for r in data_rows:
        if not r:
            continue
        i_gstin = col_map["GSTIN"]
        if i_gstin >= len(r) or not str(r[i_gstin] or "").strip():
            continue
        rows.append([r[col_map[c]] if col_map[c] is not None and col_map[c] < len(r) else None
                     for c in G2A_COLS])

    if not rows:
        raise ParseError("No invoice rows found in the B2B sheet. The file may be empty or in an unexpected format.")

    df = pd.DataFrame(rows, columns=G2A_COLS)
    df = df[df["GSTIN"].astype(str).str.strip() != ""].copy()
    for c in ["InvoiceValue", "Rate", "TaxableValue", "IGST", "CGST", "SGST", "Cess"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["GSTIN"] = df["GSTIN"].astype(str).str.strip().str.upper()
    df["InvoiceNo"] = df["InvoiceNo"].astype(str).str.strip()
    df["Period"] = df["Period"].apply(_normalize_period)
    df["MatchKey"] = df["GSTIN"] + "|" + df["InvoiceNo"].apply(_norm_inv)
    return df, generation_date


def parse_gstr2b_cdnr(path):
    """
    Returns (notes_df, generation_date_str) for the 'B2B-CDNR' sheet (credit/debit
    notes) of a GSTR-2B export. notes_df has columns CDNR_COLS + MatchKey, GSTIN
    uppercased, NoteType one of 'Credit Note'/'Debit Note' (as filed by the supplier).

    Returns an empty DataFrame (not an error) when the sheet is absent -- GSTR-2A
    exports and older GSTR-2B exports may not carry it, and a file with no notes
    for the period is a normal, not exceptional, case.
    """
    kind, wb = _open_book(path)
    sheet_names = wb.sheet_names() if kind == "xlrd" else wb.sheetnames
    if "B2B-CDNR" not in sheet_names:
        return pd.DataFrame(columns=CDNR_COLS + ["MatchKey"]), ""

    generation_date = ""
    if "Read me" in sheet_names:
        try:
            if kind == "xlrd":
                rm = wb.sheet_by_name("Read me")
                rm_rows = [rm.row_values(r) for r in range(min(15, rm.nrows))]
            else:
                rm = wb["Read me"]
                rm_rows = [list(r) for r in rm.iter_rows(min_row=1, max_row=15, values_only=True)]
            for row in rm_rows:
                for i, cell in enumerate(row):
                    if cell and "generation" in str(cell).lower():
                        for offset in (1, 2):
                            if i + offset < len(row) and row[i + offset]:
                                generation_date = str(row[i + offset])
                                break
                    if generation_date:
                        break
                if generation_date:
                    break
        except Exception:
            pass

    if kind == "xlrd":
        ws = wb.sheet_by_name("B2B-CDNR")
        all_rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        ws = wb["B2B-CDNR"]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_row_idx = None
    for i, row in enumerate(all_rows[:20]):
        cells = [_normalize_header(c) for c in row if c is not None]
        if any(c == "gstinofsupplier" for c in cells):
            header_row_idx = i
            break
    if header_row_idx is None or header_row_idx + 1 >= len(all_rows):
        raise ParseError(
            "Could not find the header row (expected a 'GSTIN of supplier' column) in the "
            "B2B-CDNR sheet. This file's layout doesn't match a recognised GSTR-2B export."
        )

    header = _combine_header_rows(all_rows[header_row_idx], all_rows[header_row_idx + 1])
    norm_header = [_normalize_header(h) for h in header]

    def find_col(*keywords):
        for i, h in enumerate(norm_header):
            if all(kw in h for kw in keywords):
                return i
        return None

    col_map = {
        "Period": find_col("period"),
        "GSTIN": find_col("gstin"),
        "SupplierName": find_col("legal"),
        "NoteNo": find_col("note", "number"),
        "NoteType": find_col("note", "type"),
        "NoteDate": find_col("note", "date"),
        "NoteValue": find_col("note", "value"),
        "PlaceOfSupply": find_col("place", "supply"),
        "RCM": find_col("reverse", "charge"),
        "TaxableValue": find_col("taxable", "value"),
        "IGST": find_col("integrated", "tax"),
        "CGST": find_col("central", "tax"),
        "SGST": find_col("state", "tax"),
        "Cess": find_col("cess"),
        "FilingDate": find_col("filing", "date"),
        "ITCAvailable": find_col("itc", "availab"),
        "Reason": find_col("reason"),
    }
    required = ["Period", "GSTIN", "NoteNo", "NoteType", "NoteDate", "NoteValue",
                "TaxableValue", "IGST", "CGST", "SGST"]
    missing = [k for k in required if col_map[k] is None]
    if missing:
        raise ParseError(
            f"Could not find expected column(s) {', '.join(missing)} in the B2B-CDNR sheet. "
            "This file's layout doesn't match a recognised GSTR-2B export."
        )

    data_rows = all_rows[header_row_idx + 2:]
    rows = []
    for r in data_rows:
        if not r:
            continue
        i_gstin = col_map["GSTIN"]
        if i_gstin >= len(r) or not str(r[i_gstin] or "").strip():
            continue
        rows.append([r[col_map[c]] if col_map[c] is not None and col_map[c] < len(r) else None
                     for c in CDNR_COLS])

    if not rows:
        return pd.DataFrame(columns=CDNR_COLS + ["MatchKey"]), generation_date

    df = pd.DataFrame(rows, columns=CDNR_COLS)
    df = df[df["GSTIN"].astype(str).str.strip() != ""].copy()
    for c in ["NoteValue", "TaxableValue", "IGST", "CGST", "SGST", "Cess"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["GSTIN"] = df["GSTIN"].astype(str).str.strip().str.upper()
    df["NoteNo"] = df["NoteNo"].astype(str).str.strip()
    df["NoteType"] = df["NoteType"].astype(str).str.strip()
    df["Period"] = df["Period"].apply(_normalize_period)
    df["MatchKey"] = df["GSTIN"] + "|" + df["NoteNo"].apply(_norm_inv)
    return df, generation_date


def _excel_serial_to_ddmmyyyy(serial):
    try:
        d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(serial))
        return d.strftime("%d/%m/%Y")
    except Exception:
        return str(serial)


def _to_pydate(val):
    """Excel date cells surface as datetime.datetime (openpyxl, date-formatted cells)
    or as a numeric day serial (xlrd, or openpyxl on plain-number cells)."""
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, (int, float)) and val > 0:
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(val))).date()
    return None


def _locate_tally_register_header(all_rows):
    """Tally register exports (Purchase/Credit Note/Debit Note) all share the same
    layout: a few rows of company letterhead, a title, a date range, then the real
    header row -- located here by content ('Date' + 'Gross Total' columns) rather
    than a fixed row number, since the letterhead varies in length."""
    for i, row in enumerate(all_rows[:40]):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if any("gross total" in c for c in cells) and any(c == "date" for c in cells):
            return i
    return None


def parse_purchase_register(path):
    """
    Returns a list of dicts ready for db.add_purchase_batch(), by locating the header
    row dynamically (looks for 'Date' + 'Gross Total' + a GSTIN-like column), then summing
    all CGST/SGST/IGST-rate-split columns into single CGST/SGST/IGST totals.
    """
    kind, wb = _open_book(path)
    sheet_names = wb.sheet_names() if kind == "xlrd" else wb.sheetnames
    target_sheet = sheet_names[0]
    for s in sheet_names:
        if "purchase" in s.lower():
            target_sheet = s
            break

    if kind == "xlrd":
        ws = wb.sheet_by_name(target_sheet)
        all_rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        ws = wb[target_sheet]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_row_idx = _locate_tally_register_header(all_rows)
    if header_row_idx is None:
        raise ParseError(
            "Could not find the header row (expected columns like 'Date' and 'Gross Total'). "
            "Please upload the Purchase Register export as-is, without removing header rows."
        )

    header = [str(c).strip() if c is not None else "" for c in all_rows[header_row_idx]]
    data_rows = all_rows[header_row_idx + 1:]

    def col_idx(*names):
        for n in names:
            for i, h in enumerate(header):
                if h.strip().lower() == n.lower():
                    return i
        return None

    i_date = col_idx("Date")
    i_particulars = col_idx("Particulars")
    i_invno = col_idx("Supplier Invoice No.", "Supplier Invoice No")
    i_invdate = col_idx("Supplier Invoice Date")
    i_gstin = col_idx("GSTIN/UIN", "GSTIN")
    i_gross = col_idx("Gross Total")

    if None in (i_date, i_invno, i_gstin, i_gross):
        raise ParseError(
            "Some required columns (Date, Supplier Invoice No., GSTIN/UIN, Gross Total) "
            "were not found by name. The file's column headers may differ from the expected export format."
        )

    buckets = [_gst_bucket(h) for h in header]
    cgst_idx = [i for i, b in enumerate(buckets) if b == "cgst"]
    sgst_idx = [i for i, b in enumerate(buckets) if b == "sgst"]
    igst_idx = [i for i, b in enumerate(buckets) if b == "igst"]

    entries = []
    for row in data_rows:
        if row is None:
            continue
        if i_date >= len(row):
            continue
        date_val = row[i_date]
        entry_date_obj = _to_pydate(date_val)
        if entry_date_obj is None:
            continue
        gstin = str(row[i_gstin]).strip().upper() if i_gstin < len(row) and row[i_gstin] else ""
        invno = str(row[i_invno]).strip() if i_invno < len(row) and row[i_invno] else ""
        if not gstin or not invno:
            continue

        cgst = sum(_num(row[i]) for i in cgst_idx if i < len(row))
        sgst = sum(_num(row[i]) for i in sgst_idx if i < len(row))
        igst = sum(_num(row[i]) for i in igst_idx if i < len(row))
        gross = _num(row[i_gross]) if i_gross < len(row) else 0
        inv_date_raw = row[i_invdate] if i_invdate is not None and i_invdate < len(row) else date_val
        inv_date_obj = _to_pydate(inv_date_raw) or entry_date_obj
        inv_date_fmt = inv_date_obj.strftime("%d/%m/%Y")
        entry_date_fmt = entry_date_obj.strftime("%d/%m/%Y")
        entry_fy = fy_utils.fy_label_for_date(entry_date_obj)
        entry_month = entry_date_obj.strftime("%Y-%m")

        entries.append({
            "entry_date": entry_date_fmt,
            "entry_fy": entry_fy,
            "entry_month": entry_month,
            "particulars": str(row[i_particulars]).strip() if i_particulars is not None and i_particulars < len(row) and row[i_particulars] else "",
            "invoice_no": invno,
            "invoice_date": inv_date_fmt,
            "gstin": gstin,
            "gross_total": round(gross, 2),
            "cgst": round(cgst, 2),
            "sgst": round(sgst, 2),
            "igst": round(igst, 2),
            "match_key": gstin + "|" + _norm_inv(invno),
        })

    if not entries:
        raise ParseError("No valid data rows were found in the Purchase Register file.")
    return entries


def parse_note_register(path, note_type):
    """
    Parses a Tally Credit Note Register or Debit Note Register export.
    note_type: 'credit' or 'debit' -- which button the user uploaded through;
    the file's own layout doesn't reliably distinguish the two (both use the
    same column set), so the caller supplies it explicitly.

    Returns a list of dicts ready for db.add_note_batch(), keyed the same way as
    parse_purchase_register (GSTIN + normalized voucher reference number), summing
    all CGST/SGST/IGST-rate-split columns into single totals.
    """
    if note_type not in ("credit", "debit"):
        raise ValueError("note_type must be 'credit' or 'debit'")

    kind, wb = _open_book(path)
    sheet_names = wb.sheet_names() if kind == "xlrd" else wb.sheetnames
    target_sheet = sheet_names[0]
    for s in sheet_names:
        if "note" in s.lower():
            target_sheet = s
            break

    if kind == "xlrd":
        ws = wb.sheet_by_name(target_sheet)
        all_rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        ws = wb[target_sheet]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_row_idx = _locate_tally_register_header(all_rows)
    label = "Credit Note" if note_type == "credit" else "Debit Note"
    if header_row_idx is None:
        raise ParseError(
            "Could not find the header row (expected columns like 'Date' and 'Gross Total'). "
            f"Please upload the {label} Register export as-is, without removing header rows."
        )

    header = [str(c).strip() if c is not None else "" for c in all_rows[header_row_idx]]
    data_rows = all_rows[header_row_idx + 1:]

    def col_idx(*names):
        for n in names:
            for i, h in enumerate(header):
                if h.strip().lower() == n.lower():
                    return i
        return None

    i_date = col_idx("Date")
    i_particulars = col_idx("Particulars")
    i_refno = col_idx("Voucher Ref. No.", "Voucher Ref. No", "Voucher Ref No.")
    i_refdate = col_idx("Voucher Ref. Date", "Voucher Ref Date")
    i_gstin = col_idx("GSTIN/UIN", "GSTIN")
    i_gross = col_idx("Gross Total")

    if None in (i_date, i_refno, i_gstin, i_gross):
        raise ParseError(
            "Some required columns (Date, Voucher Ref. No., GSTIN/UIN, Gross Total) "
            f"were not found by name. The file's column headers may differ from the expected {label} Register format."
        )

    buckets = [_gst_bucket(h) for h in header]
    cgst_idx = [i for i, b in enumerate(buckets) if b == "cgst"]
    sgst_idx = [i for i, b in enumerate(buckets) if b == "sgst"]
    igst_idx = [i for i, b in enumerate(buckets) if b == "igst"]

    entries = []
    for row in data_rows:
        if row is None:
            continue
        if i_date >= len(row):
            continue
        date_val = row[i_date]
        entry_date_obj = _to_pydate(date_val)
        if entry_date_obj is None:
            continue
        gstin = str(row[i_gstin]).strip().upper() if i_gstin < len(row) and row[i_gstin] else ""
        refno = str(row[i_refno]).strip() if i_refno < len(row) and row[i_refno] else ""
        if not gstin or not refno:
            continue

        cgst = sum(_num(row[i]) for i in cgst_idx if i < len(row))
        sgst = sum(_num(row[i]) for i in sgst_idx if i < len(row))
        igst = sum(_num(row[i]) for i in igst_idx if i < len(row))
        gross = _num(row[i_gross]) if i_gross < len(row) else 0
        ref_date_raw = row[i_refdate] if i_refdate is not None and i_refdate < len(row) else date_val
        ref_date_obj = _to_pydate(ref_date_raw) or entry_date_obj
        ref_date_fmt = ref_date_obj.strftime("%d/%m/%Y")
        entry_date_fmt = entry_date_obj.strftime("%d/%m/%Y")
        entry_fy = fy_utils.fy_label_for_date(entry_date_obj)
        entry_month = entry_date_obj.strftime("%Y-%m")

        entries.append({
            "note_type": note_type,
            "entry_date": entry_date_fmt,
            "entry_fy": entry_fy,
            "entry_month": entry_month,
            "particulars": str(row[i_particulars]).strip() if i_particulars is not None and i_particulars < len(row) and row[i_particulars] else "",
            "voucher_no": refno,
            "voucher_date": ref_date_fmt,
            "gstin": gstin,
            "gross_total": round(gross, 2),
            "cgst": round(cgst, 2),
            "sgst": round(sgst, 2),
            "igst": round(igst, 2),
            "match_key": gstin + "|" + _norm_inv(refno),
        })

    if not entries:
        raise ParseError(f"No valid data rows were found in the {label} Register file.")
    return entries
